import os
import datetime
import openpyxl
from openpyxl import Workbook
import warnings

class BPIDataProcessor:
    def __init__(self, base_path=None):
        warnings.filterwarnings('ignore', category=UserWarning, 
                                message="Cell .* is marked as a date but the serial value .* is outside the limits for dates.*")
        self.pwd = base_path or os.getcwd()
        self.input_folder = os.path.join(self.pwd, "FOR_UPLOADS")
        self.output_folder = os.path.join(self.pwd, "BPI_FOR_UPLOADS")
        os.makedirs(self.output_folder, exist_ok=True)

        self.necessary_columns = [
            'LAN', 'NAME', 'CTL4', 'PAST DUE', 'PAYOFF AMOUNT', 'PRINCIPAL', 'LPC',
            'ADA SHORTAGE', 'EMAIL', 'CONTACT NUMBER 1', 'CONTACT NUMBER 2', 
            'ENDO DATE', 'UNIT', 'DPD'
        ]

    def _get_input_file(self):
        current_date = datetime.datetime.now().strftime('%m%d%Y')
        return os.path.join(self.input_folder, f"FOR UPLOAD (NEW ENDO) {current_date}.xlsx")

    def _get_output_file(self):
        current_date = datetime.datetime.now().strftime('%m%d%Y')
        return os.path.join(self.output_folder, f"BPI AUTO CURING FOR UPLOADS {current_date}.xlsx")

    @staticmethod
    def _process_mobile_number(mobile_num):
        if not mobile_num:
            return ""
        
        mobile_num = str(mobile_num).strip().replace('-', '')
        
        if mobile_num.startswith('639') and len(mobile_num) == 12:
            return '0' + mobile_num[2:]
        
        if mobile_num.startswith('9') and len(mobile_num) == 10:
            return '0' + mobile_num 
        
        return mobile_num if mobile_num.startswith('09') else str(mobile_num)

    @staticmethod
    def _format_date(date_value):
        if date_value:
            if isinstance(date_value, datetime.datetime):
                return date_value.strftime("%m/%d/%Y")
            else:
                return str(date_value)
        return ""

    def process_file(self):
        input_file = self._get_input_file()
        output_file = self._get_output_file()

        wb_input = openpyxl.load_workbook(input_file, data_only=True, read_only=True)
        sheet = wb_input.active

        headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        col_indices = {col: headers.index(col) if col in headers else -1 for col in self.necessary_columns}

        wb_output = Workbook()
        ws = wb_output.active

        final_columns = [
            'LAN', 'CH CODE', 'NAME', 'CTL4', 'PAST DUE', 'PAYOFF AMOUNT', 'PRINCIPAL', 'LPC',
            'ADA SHORTAGE', 'EMAIL_ALS', 'MOBILE_NO_ALS', 'MOBILE_ALFES', 'LANDLINE_NO_ALFES', 
            'DATE REFERRED', 'UNIT', 'DPD'
        ]
        ws.append(final_columns)

        column_map = {
            'EMAIL': 'EMAIL_ALS',
            'CONTACT NUMBER 1': 'MOBILE_NO_ALS',
            'CONTACT NUMBER 2': 'MOBILE_ALFES',
            'ENDO DATE': 'DATE REFERRED'
        }

        for row in sheet.iter_rows(min_row=2, values_only=True):
            processed_row = []
            for col in final_columns:
                value = ""
                
                orig_col = next((k for k, v in column_map.items() if v == col), col)
                
                if orig_col in col_indices and col_indices[orig_col] != -1:
                    value = row[col_indices[orig_col]] or ""
                
                if col == 'CH CODE':
                    value = row[col_indices['LAN']] or ""
                elif col == 'MOBILE_NO_ALS':
                    value = self._process_mobile_number(value)
                elif col == 'MOBILE_ALFES':
                    value = self._process_mobile_number(value)
                elif col == 'DATE REFERRED':
                    value = self._format_date(value)
                elif col == 'LANDLINE_NO_ALFES':
                    value = ""
                elif col in ['PAST DUE', 'PAYOFF AMOUNT', 'PRINCIPAL', 'LPC', 'ADA SHORTAGE']:
                    try:
                        value = round(float(value),2)
                    except ValueError:
                        value = "0.00"
                    
                processed_row.append(str(value).strip())

            ws.append(processed_row)
        
        for col in ws.columns:
            column = col[0].column_letter
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[column].width = max_length + 2
            for cell in col:
                header_value = ws.cell(row=1, column=cell.column).value
                if header_value == 'DATE REFERRED':
                    cell.number_format = '@'
                elif header_value in ['PAST DUE', 'PAYOFF AMOUNT', 'PRINCIPAL', 'LPC', 'ADA SHORTAGE']:
                    cell.number_format = '0.00'
                    
        wb_output.save(output_file)
        print(f"Modified dataset saved to: {output_file}")

def main():
    processor = BPIDataProcessor()
    processor.process_file()

if __name__ == "__main__":
    main()