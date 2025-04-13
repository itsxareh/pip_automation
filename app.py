import streamlit as st
import pandas as pd
import os
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
import warnings
from datetime import datetime, date, time
import io
import tempfile
import shutil
import re 
import datetime
import msoffcrypto 
from supabase import create_client
from dotenv import load_dotenv
load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

warnings.filterwarnings('ignore', category=UserWarning, 
                        message="Cell .* is marked as a date but the serial value .* is outside the limits for dates.*")
class BaseProcessor:
    def __init__(self):
        self.temp_dir = tempfile.mkdtemp()
        
    def __del__(self):
        try:
            if os.path.exists(self.temp_dir):
                shutil.rmtree(self.temp_dir)
        except:
            pass
          
    def process_mobile_number(self, mobile_num):
        """Process mobile number to standardized format"""
        if not mobile_num:
            return ""
        
        mobile_num = str(mobile_num).strip().replace('-', '')
        
        if mobile_num.startswith('639') and len(mobile_num) == 12:
            return '0' + mobile_num[2:]
        
        if mobile_num.startswith('9') and len(mobile_num) == 10:
            return '0' + mobile_num 
        
        return mobile_num if mobile_num.startswith('09') else str(mobile_num)

    def format_date(self, date_value):
        if pd.isna(date_value) or date_value is None:
            return ""
            
        if isinstance(date_value, (datetime, date)):
            return date_value.strftime("%m/%d/%Y")
        
        try:
            date_obj = pd.to_datetime(date_value)
            return date_obj.strftime("%m/%d/%Y")
        except:
            return str(date_value)
    def clean_data(self, df, remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        if not isinstance(df, pd.DataFrame):
            raise ValueError(f"Expected a pandas DataFrame, but got {type(df)}: {df}")
        
        cleaned_df = df.copy()
        
        if remove_blanks: 
            cleaned_df = cleaned_df.dropna(how='all')
        if remove_duplicates:
            cleaned_df = cleaned_df.drop_duplicates()
        if trim_spaces:
            for col in cleaned_df.select_dtypes(include=['object']).columns:
                cleaned_df[col] = cleaned_df[col].str.strip()
                
        cleaned_df = cleaned_df.replace(r'^\s*$', pd.NA, regex=True)
        return cleaned_df
        
    def clean_only(self, file_content, preview_only=False, 
               remove_duplicates=False, remove_blanks=False, trim_spaces=False, file_name=None):
        try:
            df = pd.read_excel(io.BytesIO(file_content))

            sanitized_headers = [re.sub(r'[^A-Za-z0-9_]', '_', str(col)) for col in df.columns]
            df.columns = sanitized_headers

            cleaned_df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)

            if preview_only:
                return cleaned_df

            if file_name:
                base_name = os.path.splitext(os.path.basename(file_name))[0]
                output_filename = f"{base_name}.xlsx"
            else:
                output_filename = f"CLEANED_DATA.xlsx"

            output_path = os.path.join(self.temp_dir, output_filename)

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                cleaned_df.to_excel(writer, index=False, sheet_name='Sheet1')

                worksheet = writer.sheets['Sheet1']
                for i, col in enumerate(cleaned_df.columns):
                    try:
                        max_len_in_column = cleaned_df[col].astype(str).map(len).max()
                        max_length = max(max_len_in_column, len(str(col))) + 2
                    except:
                        max_length = 15
                    col_letter = get_column_letter(i + 1)
                    worksheet.column_dimensions[col_letter].width = max_length

            with open(output_path, 'rb') as f:
                output_binary = f.read()

            return cleaned_df, output_binary, output_filename

        except Exception as e:
            st.error(f"Error cleaning file: {str(e)}")
            raise
class BPIProcessor(BaseProcessor):
    
    def setup_directories(self, automation_type):
        """Create necessary directories based on automation type"""
        directories = {
            'updates': ["FOR_UPDATES", "BPI_FOR_UPDATES"],
            'uploads': ["FOR_UPLOADS", "BPI_FOR_UPLOADS"],
            'cured_list': ["CURED_LIST", "BPI_FOR_REMARKS", "BPI_FOR_PAYMENTS", "BPI_FOR_OTHERS"]
        }
        
        dirs_to_create = directories.get(automation_type, [])
        created_dirs = {}
        
        for dir_name in dirs_to_create:
            dir_path = os.path.join(self.temp_dir, dir_name)
            os.makedirs(dir_path, exist_ok=True)
            created_dirs[dir_name] = dir_path
            
        return created_dirs
        
    def process_updates_or_uploads(self, file_content, automation_type, preview_only=False,
                                   remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        try:
            df = pd.read_excel(io.BytesIO(file_content))
            df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)
            
            if preview_only:
                return df
                
            current_date = datetime.now().strftime('%m%d%Y')
            
            if automation_type == 'updates':
                output_filename = f"BPI AUTO CURING FOR UPDATES {current_date}.xlsx"
                input_filename = f"FOR UPDATE {current_date}.xlsx"
                dirs = self.setup_directories('updates')
                folder_key = 'BPI_FOR_UPDATES'
                input_folder_key = 'FOR_UPDATES'
            else: 
                output_filename = f"BPI AUTO CURING FOR UPLOADS {current_date}.xlsx"
                input_filename = f"FOR UPLOAD (NEW ENDO) {current_date}.xlsx"
                dirs = self.setup_directories('uploads')
                folder_key = 'BPI_FOR_UPLOADS'
                input_folder_key = 'FOR_UPLOADS'
            
            input_path = os.path.join(dirs[input_folder_key], input_filename)
            with open(input_path, 'wb') as f:
                f.write(file_content)
                
            column_map = {
                'EMAIL': 'EMAIL_ALS',
                'CONTACT NUMBER 1': 'MOBILE_NO_ALS',
                'CONTACT NUMBER 2': 'MOBILE_ALFES',
                'ENDO DATE': 'DATE REFERRED'
            }
            
            result_df = pd.DataFrame()
            
            for col in ['LAN', 'NAME', 'CTL4', 'PAST DUE', 'PAYOFF AMOUNT', 'PRINCIPAL', 'LPC', 
                        'ADA SHORTAGE', 'UNIT', 'DPD']:
                if col in df.columns:
                    result_df[col] = df[col].fillna("")
            
            result_df.insert(1, 'CH CODE', result_df['LAN'])
            
            for orig_col, new_col in column_map.items():
                if orig_col in df.columns:
                    if orig_col == 'CONTACT NUMBER 1' or orig_col == 'CONTACT NUMBER 2':
                        result_df[new_col] = df[orig_col].apply(lambda x: "" if pd.isna(x) else self.process_mobile_number(x))
                    elif orig_col == 'ENDO DATE':
                        result_df[new_col] = df[orig_col].apply(lambda x: self.format_date(x) if pd.notnull(x) else "")
                    else:
                        result_df[new_col] = df[orig_col].fillna("")
                else:
                    result_df[new_col] = ""
            
            result_df['LANDLINE_NO_ALFES'] = ""
            
            numeric_cols = ['PAST DUE', 'PAYOFF AMOUNT', 'PRINCIPAL', 'LPC', 'ADA SHORTAGE']
            for col in numeric_cols:
                if col in result_df.columns:
                    result_df[col] = pd.to_numeric(result_df[col], errors='coerce').fillna(0).round(2)
                    
            final_columns = [
                'LAN', 'CH CODE', 'NAME', 'CTL4', 'PAST DUE', 'PAYOFF AMOUNT', 'PRINCIPAL', 'LPC',
                'ADA SHORTAGE', 'EMAIL_ALS', 'MOBILE_NO_ALS', 'MOBILE_ALFES', 'LANDLINE_NO_ALFES', 
                'DATE REFERRED', 'UNIT', 'DPD'
            ]
            
            for col in final_columns:
                if col not in result_df.columns:
                    result_df[col] = ""
                    
            result_df = result_df[final_columns]
            
            output_path = os.path.join(dirs[folder_key], output_filename)
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Sheet1')
                
                worksheet = writer.sheets['Sheet1']
                for i, col in enumerate(final_columns):
                    max_length = max(
                        result_df[col].astype(str).map(len).max(),
                        len(col)
                    ) + 2
                    col_letter = chr(65 + i) 
                    worksheet.column_dimensions[col_letter].width = max_length
                    
                    if col in numeric_cols:
                        for row in range(2, len(result_df) + 2):
                            cell = worksheet[f"{col_letter}{row}"]
                            cell.number_format = '0.00'
                    
                    if col == 'DATE REFERRED':
                        for row in range(2, len(result_df) + 2):
                            cell = worksheet[f"{col_letter}{row}"]
                            value = cell.value
                            if value:
                                try: 
                                    cell.value = pd.to_datetime(value).strftime("%m/%d/%Y")
                                    cell.number_format = '@'
                                except:
                                    pass
            
            with open(output_path, 'rb') as f:
                output_binary = f.read()
                
            return result_df, output_binary, output_filename
            
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
            raise

    def process_updates(self, file_content, preview_only=False,
                        remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        return self.process_updates_or_uploads(file_content, 'updates', preview_only,
                                               remove_duplicates, remove_blanks, trim_spaces)
        
    def process_uploads(self, file_content, preview_only=False,
                        remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        return self.process_updates_or_uploads(file_content, 'uploads', preview_only,
                                               remove_duplicates, remove_blanks, trim_spaces)
    
    def process_cured_list(self, file_content, preview_only=False,
                           remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_input:
            temp_input.write(file_content)
            temp_input_path = temp_input.name
            
        try:
            
            df = pd.read_excel(temp_input_path)
            df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)
            
            if preview_only:
                return df
                
            current_date = datetime.now().strftime('%m%d%Y')
            dirs = self.setup_directories('cured_list')
            
            input_file = os.path.join(dirs["CURED_LIST"], f"CURED LIST {current_date}.xlsx")
            shutil.copy(temp_input_path, input_file)
            
            remarks_filename = f"BPI AUTOCURING REMARKS {current_date}.xlsx"
            others_filename = f"BPI AUTOCURING RESHUFFLE {current_date}.xlsx"
            payments_filename = f"BPI AUTOCURING PAYMENT {current_date}.xlsx"
            
            remarks_path = os.path.join(dirs["BPI_FOR_REMARKS"], remarks_filename)
            others_path = os.path.join(dirs["BPI_FOR_OTHERS"], others_filename)
            payments_path = os.path.join(dirs["BPI_FOR_PAYMENTS"], payments_filename)
            
            try:
                source_wb = openpyxl.load_workbook(input_file)
            except FileNotFoundError:
                print(f"Error: The file '{input_file}' was not found.")
                return
            
            dest_wb = openpyxl.Workbook()
            dest_ws = dest_wb.active
            
            headers = ["LAN", "Action Status", "Remark Date", "PTP Date", "Reason For Default", 
                    "Field Visit Date", "Remark", "Next Call Date", "PTP Amount", "Claim Paid Amount", 
                    "Remark By", "Phone No.", "Relation", "Claim Paid Date"]
            
            for col, header in enumerate(headers, 1):
                dest_ws.cell(row=1, column=col).value = header
            
            ws = source_wb.active
            if ws.max_column < 43:
                raise ValueError("File doesn't have the expected number of columns")
            
            current_row = 2
            total_rows = 0
            last_row = ws.max_row
            
            barcode_lookup = {}
            for row in range(2, last_row + 1):
                barcode = ws.cell(row=row, column=1).value
                if barcode:
                    barcode_lookup[barcode] = {
                        'date': ws.cell(row=row, column=3).value, 
                        'amount': ws.cell(row=row, column=4).value, 
                        'collector': ws.cell(row=row, column=2).value,  
                        'phone1': ws.cell(row=row, column=42).value, 
                        'phone2': ws.cell(row=row, column=43).value, 
                    }
            
            nego_rows = []
            for row in range(2, last_row + 1):
                if (ws.cell(row=row, column=2).value != "SPMADRID" and 
                    (ws.cell(row=row, column=8).value is None or "PTP" not in str(ws.cell(row=row, column=8).value))):
                    nego_rows.append(row)
            
            if nego_rows:
                visible_count = len(nego_rows)
                
                for i, row_idx in enumerate(nego_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PTP NEW - CALL OUTS_PASTDUE"
                
                current_row += visible_count
                for i, row_idx in enumerate(nego_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PTP FF UP - CLIENT ANSWERED AND WILL SETTLE"
                
                current_row += visible_count
                
                for i, row_idx in enumerate(nego_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PAYMENT - CURED"
                
                current_row += visible_count
                
                total_rows += (visible_count * 3)
            
            ptp_rows = []
            for row in range(2, last_row + 1):
                if (ws.cell(row=row, column=2).value != "SPMADRID" and 
                    ws.cell(row=row, column=8).value is not None and 
                    "PTP" in str(ws.cell(row=row, column=8).value)):
                    ptp_rows.append(row)
            
            if ptp_rows:
                visible_count = len(ptp_rows)
                
                for i, row_idx in enumerate(ptp_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PTP FF UP - CLIENT ANSWERED AND WILL SETTLE"
                
                current_row += visible_count
                
                for i, row_idx in enumerate(ptp_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PAYMENT - CURED"
                
                current_row += visible_count
                
                total_rows += (visible_count * 2)
            
            spmadrid_rows = []
            for row in range(2, last_row + 1):
                if ws.cell(row=row, column=2).value == "SPMADRID":
                    spmadrid_rows.append(row)
            
            if spmadrid_rows:
                visible_count = len(spmadrid_rows)
                
                for i, row_idx in enumerate(spmadrid_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PTP NEW - CURED_GHOST"
                
                current_row += visible_count
                
                for i, row_idx in enumerate(spmadrid_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PAYMENT - CURED"
                
                current_row += visible_count
                
                total_rows += (visible_count * 2)
            
            final_row_count = total_rows + 1
            
            for row in range(2, final_row_count + 1):
                barcode = dest_ws.cell(row=row, column=1).value
                action_status = dest_ws.cell(row=row, column=2).value
                
                source_data = barcode_lookup.get(barcode, {})
                source_date = source_data.get('date')
                source_amount = source_data.get('amount')
                source_collector = source_data.get('collector')
                source_phone1 = source_data.get('phone1')
                source_phone2 = source_data.get('phone2')
                
                if source_date:
                    try:
                        if hasattr(source_date, 'strftime'): 
                            base_date = source_date
                        else:
                            try:
                                base_date = datetime.strptime(str(source_date), "%Y-%m-%d %H:%M:%S")
                            except:
                                try:
                                    base_date = datetime.strptime(str(source_date), "%Y-%m-%d")
                                except:
                                    base_date = datetime.now()
                    except:
                        base_date = datetime.now()
                    
                    if "PTP NEW" in action_status:
                        time_to_add = time(14, 40, 0)
                    elif "PTP FF" in action_status:
                        time_to_add = time(14, 50, 0)
                    elif "CURED" in action_status:
                        time_to_add = time(15, 0, 0)
                    else:
                        time_to_add = time(0, 0, 0)
                    
                    if not hasattr(base_date, 'time'):
                        base_date = datetime.combine(base_date, time(0, 0, 0))
                    
                    result_date = datetime.combine(base_date.date(), time_to_add)
                    
                    formatted_date = result_date.strftime("%m/%d/%Y %I:%M:%S %p")
                    dest_ws.cell(row=row, column=3).value = formatted_date
                    
                    formatted_date = result_date.strftime("%m/%d/%Y")
                    dest_ws.cell(row=row, column=4).value = formatted_date
                    
                    dest_ws.cell(row=row, column=3).number_format = '@'
                    dest_ws.cell(row=row, column=4).number_format = '@'
                else:
                    dest_ws.cell(row=row, column=3).value = ""
                    dest_ws.cell(row=row, column=4).value = ""
                
                phone_no = ""
                if "PAYMENT" not in action_status:
                    phone_no = dest_ws.cell(row=row, column=12).value
                
                if "PTP NEW" in action_status:
                    phone_value = source_phone1 if source_phone1 else source_phone2
                    remark_text = f"1_{self.process_mobile_number(phone_value)} - PTP NEW"
                elif "PTP FF" in action_status:
                    phone_value = source_phone1 if source_phone1 else source_phone2
                    remark_text = f"{self.process_mobile_number(phone_value)} - FPTP"
                elif "PAYMENT" in action_status:
                    remark_text = "CURED - CONFIRM VIA SELECTIVE LIST"
                else:
                    remark_text = ""
                
                dest_ws.cell(row=row, column=7).value = remark_text
                
                if "PAYMENT" in action_status:
                    dest_ws.cell(row=row, column=9).value = ""
                else:
                    dest_ws.cell(row=row, column=9).value = source_amount
                
                if "PAYMENT" in action_status:
                    dest_ws.cell(row=row, column=10).value = source_amount
                else:
                    dest_ws.cell(row=row, column=10).value = ""
                
                dest_ws.cell(row=row, column=11).value = source_collector
                
                if "PAYMENT" in action_status:
                    dest_ws.cell(row=row, column=12).value = ""
                else:
                    if source_phone1 and source_phone1 != "":
                        dest_ws.cell(row=row, column=12).value = source_phone1
                    else:
                        dest_ws.cell(row=row, column=12).value = source_phone2
                
                if "PAYMENT" in action_status and source_date:
                    if isinstance(source_date, datetime):
                        formatted_paid_date = source_date.strftime("%m/%d/%Y")
                    elif isinstance(source_date, date):
                        formatted_paid_date = source_date.strftime("%m/%d/%Y")
                    else:
                        try:
                            date_obj = datetime.strptime(str(source_date), "%Y-%m-%d %H:%M:%S")
                            formatted_paid_date = date_obj.strftime("%m/%d/%Y")
                        except:
                            try:
                                date_obj = datetime.strptime(str(source_date), "%Y-%m-%d")
                                formatted_paid_date = date_obj.strftime("%m/%d/%Y")
                            except:
                                formatted_paid_date = ""
                    dest_ws.cell(row=row, column=14).value = formatted_paid_date
                else:
                    dest_ws.cell(row=row, column=14).value = ""
            
            for row in range(2, final_row_count + 1):
                action_status = dest_ws.cell(row=row, column=2).value
                phone_no = dest_ws.cell(row=row, column=12).value
                
                if "PTP NEW" in action_status and phone_no:
                    dest_ws.cell(row=row, column=7).value = f"1_{phone_no} - PTP NEW"
                elif "PTP FF" in action_status and phone_no:
                    dest_ws.cell(row=row, column=7).value = f"{phone_no} - FPTP"
            
            for column in dest_ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                dest_ws.column_dimensions[column_letter].width = adjusted_width

            for row_idx in range(2, dest_ws.max_row + 1):  
                for col_idx in [3, 4, 14]: 
                    cell = dest_ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        cell_value_str = str(cell.value)
                        cell.value = cell_value_str
                        cell.number_format = '@'

            dest_wb.save(remarks_path)
            
            others_wb = openpyxl.Workbook()
            others_ws = others_wb.active
            
            others_ws.cell(row=1, column=1).value = ws.cell(row=1, column=1).value 
            others_ws.cell(row=1, column=2).value = "REMARK BY" 
            
            for row in range(2, last_row + 1):
                others_ws.cell(row=row, column=1).value = ws.cell(row=row, column=1).value

                reference_value = ws.cell(row=row, column=1).value 
                
                for cured_row in range(2, ws.max_row + 1):  
                    if ws.cell(row=cured_row, column=1).value == reference_value: 
                        others_ws.cell(row=row, column=2).value = ws.cell(row=cured_row, column=2).value 
                        break

            others_wb.save(others_path)
            
            payments_wb = openpyxl.Workbook()
            payments_ws = payments_wb.active
            payments_ws.cell(row=1, column=1).value = "LAN"
            payments_ws.cell(row=1, column=2).value = "ACCOUNT NUMBER"
            payments_ws.cell(row=1, column=3).value = "NAME"
            payments_ws.cell(row=1, column=4).value = "CARD NUMBER"
            payments_ws.cell(row=1, column=5).value = "PAYMENT AMOUNT"
            payments_ws.cell(row=1, column=6).value = "PAYMENT DATE"
            
            for row in range(2, last_row + 1):
                payments_ws.cell(row=row, column=1).value = ws.cell(row=row, column=17).value if ws.cell(row=row, column=17).value else ""
                payments_ws.cell(row=row, column=3).value = ws.cell(row=row, column=18).value if ws.cell(row=row, column=18).value else ""
                payments_ws.cell(row=row, column=5).value = ws.cell(row=row, column=4).value if ws.cell(row=row, column=4).value else ""
                date_value = ws.cell(row=row, column=3).value
                if date_value:
                    if isinstance(date_value, datetime):
                        formatted_date = date_value.strftime("%m/%d/%Y")
                    else:
                        formatted_date = str(date_value)
                    payments_ws.cell(row=row, column=6).value = formatted_date
            
            for row in range(2, last_row + 1):
                payments_ws.cell(row=row, column=6).number_format = "@"
            
            for column in payments_ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                payments_ws.column_dimensions[column_letter].width = adjusted_width
            
            payments_wb.save(payments_path)
            
            remarks_df = pd.read_excel(remarks_path)
            others_df = pd.read_excel(others_path)
            payments_df = pd.read_excel(payments_path)
            
            with open(remarks_path, 'rb') as f:
                remarks_binary = f.read()
            with open(others_path, 'rb') as f:
                others_binary = f.read()
            with open(payments_path, 'rb') as f:
                payments_binary = f.read()
                
            os.unlink(temp_input_path)
            
            return {
                'remarks_df': remarks_df, 
                'others_df': others_df, 
                'payments_df': payments_df,
                'remarks_binary': remarks_binary,
                'others_binary': others_binary,
                'payments_binary': payments_binary,
                'remarks_filename': remarks_filename,
                'others_filename': others_filename,
                'payments_filename': payments_filename
            }
        finally:
            if os.path.exists(temp_input_path):
                os.unlink(temp_input_path)

class ROBBikeProcessor(BaseProcessor):
    def process_daily_remark(self, file_content, preview_only=False,
                    remove_duplicates=False, remove_blanks=False, trim_spaces=False, report_date=None):
        try:
            df = pd.read_excel(io.BytesIO(file_content))
            
            df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)
            df = df.iloc[:-1].reset_index(drop=True)
            
            if 'Time' in df.columns:
                df = df.sort_values(by='Time', ascending=False)
            
            if 'Account No.' in df.columns and 'Status' in df.columns:
                df['COMBINED_KEY'] = df['Account No.'].astype(str) + '_' + df['Status'].astype(str)
                
                df = df.drop_duplicates(subset=['COMBINED_KEY'])
                df = df.drop(columns=['COMBINED_KEY'])  
                
            if 'PTP Amount' in df.columns and 'Status' in df.columns:
                voluntary_surrender_rows = df[df['Status'] == 'PTP - VOLUNTARY SURRENDER']

                invalid_amount_rows = voluntary_surrender_rows[
                    (voluntary_surrender_rows['PTP Amount'].isna()) |
                    (voluntary_surrender_rows['PTP Amount'] == 0)
                ]

                if not invalid_amount_rows.empty:
                    st.warning(f"Found {len(invalid_amount_rows)} row(s) with 'PTP - VOLUNTARY SURRENDER' but 0 or missing 'PTP Amount'.")
                    st.dataframe(invalid_amount_rows, use_container_width=True)
                    
            if preview_only:
                return df, None, None
            
            output_template = "DAILY MONITORING PTP, DEPO & REPO REPORT TEMPLATE.xlsx"
            sheet1 = "MONITORING"
            sheet2 = "PTP"
            sheet3 = "REPO"
            sheet4 = "DEPO"
            sheet5 = "EOD"
            
            monitoring_columns = ['Account Name', 'Account Number', 'Principal', 'EndoDate', 'Stores', 
                                'Cluster', 'DaysPastDue', 'Field Status', 'Field Substatus', 
                                'Status', 'subStatus', 'Notes', 'BarcodeDate', 'PTP Amount', 'PTP Date']
            monitoring_df = pd.DataFrame(columns=monitoring_columns)
            
            ptp_columns = ['Account Name', 'AccountNumber', 'Status', 'subStatus', 'Amount', 
                        'StartDate', 'Notes', 'ResultDate', 'EndoDate']
            ptp_df = pd.DataFrame(columns=ptp_columns)
            
            if 'Debtor' in df.columns:
                monitoring_df['Account Name'] = df['Debtor']
            
            if 'Account No.' in df.columns:
                monitoring_df['Account Number'] = df['Account No.']
            
            if 'Balance' in df.columns:
                monitoring_df['Principal'] = df['Balance']
            
            if 'DPD' in df.columns:
                monitoring_df['DaysPastDue'] = df['DPD']
            
            if 'Status' in df.columns:
                status_parts = df['Status'].str.split('-', n=1)
                monitoring_df['Status'] = status_parts.str[0].str.strip()
                
                monitoring_df['subStatus'] = status_parts.str[1].str.strip().where(status_parts.str.len() > 1, "")
            
            if 'Remark' in df.columns:
                monitoring_df['Notes'] = df['Remark']
            
            if 'Date' in df.columns:
                monitoring_df['BarcodeDate'] = pd.to_datetime(df['Date']).dt.strftime('%m/%d/%Y')
            
            if 'PTP Amount' in df.columns:
                monitoring_df['PTP Amount'] = df['PTP Amount']
            
            if 'PTP Date' in df.columns:
                monitoring_df['PTP Date'] = pd.to_datetime(df['PTP Date']).dt.strftime('%m/%d/%Y')
            
            if 'Account No.' in df.columns:
                account_numbers = [str(int(acc)) for acc in df['Account No.'].dropna().unique().tolist()]
                
                dataset_response = supabase.table('rob_bike_dataset').select('*').in_('account_number', account_numbers).execute()
                
                if hasattr(dataset_response, 'data') and dataset_response.data:
                    dataset_df = pd.DataFrame(dataset_response.data)
                    monitoring_df['Account Number'] = monitoring_df['Account Number'].apply(lambda x: str(int(float(x))) if pd.notnull(x) else '')
                    
                    account_data_map = {}
                    chcode_list = []
                    
                    for _, row in dataset_df.iterrows():
                        account_no = str(row['account_number']).strip()
                        chcode = row.get('chcode', '')
                        
                        if chcode:
                            chcode_list.append(chcode)
                            
                        account_data_map[account_no] = {
                            'ChCode': chcode,
                            'EndoDate': row.get('endo_date', ''),
                            'Stores': row.get('stores', ''),
                            'Cluster': row.get('cluster', '')
                        }
                    
                    if chcode_list:
                        try:
                            field_results_response = supabase.table('rob_bike_field_result').select('*').in_('chcode', chcode_list).execute()
                            
                            if hasattr(field_results_response, 'data') and field_results_response.data:
                                field_results_df = pd.DataFrame(field_results_response.data)
                                
                                if 'inserted_date' in field_results_df.columns:
                                    field_results_df['inserted_date'] = pd.to_datetime(field_results_df['inserted_date'])
                                
                                latest_status_map = {}
                                
                                if 'inserted_date' in field_results_df.columns:
                                    for chcode, group in field_results_df.groupby('chcode'):
                                        latest_row = group.sort_values('inserted_date', ascending=False).iloc[0]
                                        
                                        latest_status_map[chcode] = {
                                            'Field_Status': latest_row.get('status', ''),
                                            'Field_Substatus': latest_row.get('substatus', ''),
                                        }
                                    
                                    for account_no, data in account_data_map.items():
                                        chcode = data['ChCode']
                                        if chcode in latest_status_map:
                                            account_data_map[account_no].update({
                                                'Field_Status': latest_status_map[chcode]['Field_Status'],
                                                'Field_Substatus': latest_status_map[chcode]['Field_Substatus'],
                                            })
                                        else:
                                            account_data_map[account_no].update({
                                                'Field_Status': '',
                                                'Field_Substatus': '',
                                            })
                        except Exception as e:
                            st.error(f"Error fetching field results: {str(e)}")
                    
                    monitoring_df['EndoDate'] = monitoring_df['Account Number'].map(
                        lambda acc_no: account_data_map.get(acc_no, {}).get('EndoDate', ''))
                    
                    monitoring_df['Stores'] = monitoring_df['Account Number'].map(
                        lambda acc_no: account_data_map.get(acc_no, {}).get('Stores', ''))
                    
                    monitoring_df['Cluster'] = monitoring_df['Account Number'].map(
                        lambda acc_no: account_data_map.get(acc_no, {}).get('Cluster', ''))
                    
                    monitoring_df['Field Status'] = monitoring_df['Account Number'].map(
                        lambda acc_no: account_data_map.get(acc_no, {}).get('Field_Status', ''))
                    
                    monitoring_df['Field Substatus'] = monitoring_df['Account Number'].map(
                        lambda acc_no: account_data_map.get(acc_no, {}).get('Field_Substatus', ''))
                    
            ptp_data = df[df['Status'].str.contains('PTP', case=False, na=False)].copy() if 'Status' in df.columns else pd.DataFrame()
            
            if not ptp_data.empty:
                if 'Debtor' in ptp_data.columns:
                    ptp_df['Account Name'] = ptp_data['Debtor']
                
                if 'Account No.' in ptp_data.columns:
                    ptp_df['AccountNumber'] = ptp_data['Account No.']
                
                if 'Status' in ptp_data.columns:
                    status_parts = ptp_data['Status'].str.split('-', n=1)
                    ptp_df['Status'] = status_parts.str[0].str.strip()
                    ptp_df['subStatus'] = status_parts.str[1].str.strip().where(status_parts.str.len() > 1, "")
                
                if 'PTP Amount' in ptp_data.columns:
                    ptp_df['Amount'] = ptp_data['PTP Amount']
                
                if 'PTP Date' in ptp_data.columns:
                    ptp_df['StartDate'] = pd.to_datetime(ptp_data['PTP Date']).dt.strftime('%Y-%m-%d')
                
                if 'Remark' in ptp_data.columns:
                    ptp_df['Notes'] = ptp_data['Remark']
                
                if 'Time' in ptp_data.columns:
                    ptp_df['ResultDate'] = pd.to_datetime(ptp_data['Time']).dt.strftime('%m/%d/%Y %H:%M')
                
                if 'Account No.' in ptp_data.columns and 'account_data_map' in locals():
                    ptp_df['AccountNumber'] = ptp_df['AccountNumber'].apply(lambda x: str(int(float(x))) if pd.notnull(x) else '')
                    ptp_df['EndoDate'] = ptp_df['AccountNumber'].map(
                        lambda acc_no: account_data_map.get(acc_no, {}).get('EndoDate', ''))
                    ptp_df['EndoDate'] = pd.to_datetime(ptp_df['EndoDate']).dt.strftime('%m/%d/%Y')
            
            template_path = os.path.join(os.path.dirname(__file__), output_template)
            
            output_buffer = io.BytesIO()
            
            if os.path.exists(template_path):
                
                with open(template_path, 'rb') as template_file:
                    template_copy = io.BytesIO(template_file.read()) 

                    template_wb = load_workbook(template_copy)

                    def append_df_to_sheet(sheet_name, df):
                        if sheet_name in template_wb.sheetnames:
                            sheet = template_wb[sheet_name]
                            start_row = sheet.max_row + 1
                            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row):
                                for c_idx, value in enumerate(row, 1):
                                    sheet.cell(row=r_idx, column=c_idx).value = value

                    append_df_to_sheet(sheet1, monitoring_df)
                    append_df_to_sheet(sheet2, ptp_df)
                    append_df_to_sheet(sheet5, monitoring_df)

                    template_wb.save(output_buffer)
                
            else:
                with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                    monitoring_df.to_excel(writer, sheet_name=sheet1, index=False)
                    ptp_df.to_excel(writer, sheet_name=sheet2, index=False)
                    monitoring_df.to_excel(writer, sheet_name=sheet5, index=False)

                    workbook = writer.book
                    
                    workbook.create_sheet(title=sheet3)
                    workbook.create_sheet(title=sheet4)
                    
                    def format_sheet(sheet_name, df=None):
                        sheet = writer.sheets.get(sheet_name) or workbook[sheet_name]
                        
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'),
                        )
                        
                        if df is not None: 
                            for col_idx, col in enumerate(df.columns, 1):
                                max_length = max(
                                    df[col].astype(str).map(len).max(),
                                    len(str(col))
                                )
                                adjusted_width = max_length + 2
                                sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
                            
                            for row in sheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                                for cell in row:
                                    cell.border = thin_border
                                    
                    format_sheet(sheet1, monitoring_df)
                    format_sheet(sheet2, ptp_df)
                    format_sheet(sheet3)
                    format_sheet(sheet4)
                    format_sheet(sheet5, monitoring_df)
                    
            output_buffer.seek(0)
            
            if not report_date:
                report_date = datetime.now()

            date_str = report_date.strftime("%d%b%Y").upper()
            
            output_filename = f"DAILY MONITORING PTP, DEPO & REPO REPORT as of {date_str}.xls"
            
            return monitoring_df, output_buffer.getvalue(), output_filename
        
        except Exception as e:
            st.error(f"Error processing daily remark: {str(e)}")
            import traceback
            return None, None, None

class NoProcessor(BaseProcessor):
    pass

CAMPAIGN_CONFIG = {
    "No Campaign": {
        "automation_options": ["Data Clean"],
        "automation_map": {
            "Data Clean": "clean_only",
        },
        "processor": NoProcessor    
    },
    "BPI": {
        "automation_options": ["Data Clean", "Uploads", "Updates", "Cured List"],
        "automation_map": {
            "Data Clean": "clean_only",
            "Uploads": "process_uploads",
            "Updates": "process_updates",
            "Cured List": "process_cured_list"
        },
        "processor": BPIProcessor
    },
    "ROB Bike": {
        "automation_options": ["Data Clean", "Daily Remark Report"],
        "automation_map": {
            "Data Clean": "clean_only",
            "Daily Remark Report": "process_daily_remark",
        },
        "processor": ROBBikeProcessor
    }
}

def main():
    st.set_page_config(page_title="Automation Tool", layout="wide")
    st.title("Automation Tool")
    st.markdown("Transform Files into CMS Format")

    campaign = st.sidebar.selectbox("Select Campaign", ["No Campaign","BPI", "ROB Bike"], index=0)
    config = CAMPAIGN_CONFIG[campaign]
    processor = config["processor"]()
    automation_map = config["automation_map"]
    automation_options = config["automation_options"]

    st.sidebar.header("Settings")
    automation_type = st.sidebar.selectbox("Select Automation Type", automation_options, key=f"{campaign}_automation_type")

    preview = st.sidebar.checkbox("Preview file before processing", value=True, key=f"{campaign}_preview")
    st.markdown("""
            <style>
            div[data-testid="stFileUploaderDropzoneInstructions"] {
                display: none;
            }
            section[data-testid="stFileUploaderDropzone"] {
                padding: 0px;
                margin: 0px;
            }
            button[data-testid="stBaseButton-secondary"] {
                width: 100%;
            }
            </style>
        """, unsafe_allow_html=True)
    uploaded_file = st.sidebar.file_uploader(
        "Upload File", 
        type=["xlsx", "xls"], 
        key=f"{campaign}_file_uploader"
    )
    
    if campaign == "ROB Bike" and automation_type == "Daily Remark Report":
        report_date = st.sidebar.date_input('Date Report', format="YYYY/MM/DD") 
        with st.sidebar.expander("Upload Other File", expanded=False):
            upload_field_result = st.file_uploader(
                "Field Result",
                type=["xlsx", "xls"],
                key=f"{campaign}_field_result"
            )
            upload_dataset = st.file_uploader(
                "Dataset",
                type=["xlsx", "xls"],
                key=f"{campaign}_dataset"
            )
        if upload_field_result:
            TABLE_NAME = 'rob_bike_field_result'
            
            try:
                xls = pd.ExcelFile(upload_field_result)
                sheet_name = next((s for s in xls.sheet_names if s.lower() == 'result'), None)

                if sheet_name:
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    df_clean = df.replace({np.nan: 0})
                else:
                    st.error("Sheet named 'RESULT' not found in the uploaded file.")
                
                df_filtered = df_clean[(df_clean['status'] != 'CANCEL') & (df_clean['bank'] == 'ROB MOTOR LOAN')].copy()
                
                if 'chcode' in df_filtered.columns and 'status' in df_filtered.columns and 'SUB STATUS' in df_filtered.columns and 'DATE' in df_filtered.columns and 'TIME' in df_filtered.columns:
                    df_extracted = df_filtered[['chcode', 'status', 'SUB STATUS', 'DATE', 'TIME']].copy()
                    
                    df_extracted = df_extracted.rename(columns={
                        'SUB STATUS': 'substatus',
                        'DATE': 'date',
                        'TIME': 'time'
                    })
                    
                    df_extracted.loc[:, 'time'] = df_extracted['time'].astype(str).replace('NaT', '')
            
                    try:
                        temp_dates = pd.to_datetime(df_extracted['date'], errors='coerce')
                        df_extracted.loc[:, 'date'] = temp_dates.astype(str).str.split(' ').str[0]
                        df_extracted.loc[:, 'date'] = df_extracted['date'].replace('NaT', '')
                    except:
                        df_extracted.loc[:, 'date'] = df_extracted['date'].astype(str).replace('NaT', '')

                    df_extracted['inserted_date'] = pd.to_datetime(
                        df_extracted['date'].astype(str) + ' ' + df_extracted['time'].astype(str), 
                        errors='coerce'
                    )

                    df_extracted['inserted_date'] = df_extracted['inserted_date'].astype(str).replace('NaT', None)

                    st.subheader("Extracted Field Result Data:")
                    st.dataframe(df_extracted)
                    
                    button_placeholder = st.empty()
                    status_placeholder = st.empty()
                    
                    upload_button = button_placeholder.button("Upload to Database", key="upload_button")
                    
                    if upload_button:
                        button_placeholder.button("Processing...", disabled=True, key="processing_button")
                        
                        try:
                            existing_records_response = supabase.table(TABLE_NAME).select("chcode, status, substatus").execute()
                            
                            if hasattr(existing_records_response, 'data'):
                                existing_records = existing_records_response.data
                                existing_df = pd.DataFrame(existing_records) if existing_records else pd.DataFrame()
                            else:
                                existing_df = pd.DataFrame()
                            
                            df_to_upload = df_extracted.copy()
                            for col in df_to_upload.columns:
                                if pd.api.types.is_datetime64_any_dtype(df_to_upload[col]):
                                    df_to_upload[col] = df_to_upload[col].dt.strftime('%Y-%m-%d')

                            df_to_upload = df_to_upload.astype(object).where(pd.notnull(df_to_upload), None)
                            records_to_insert = df_to_upload.to_dict(orient="records")
                            
                            filtered_records = []
                            total_records = len(records_to_insert)
                            duplicate_count = 0
                            
                            progress_bar = st.progress(0)
                            status_text = status_placeholder.empty()
                            
                            for i, record in enumerate(records_to_insert):
                                if not existing_df.empty:
                                    matching = existing_df[
                                        (existing_df['chcode'] == record['chcode']) & 
                                        (existing_df['status'] == record['status']) & 
                                        (existing_df['substatus'] == record['substatus'])
                                    ]
                                    
                                    if matching.empty:
                                        filtered_records.append(record)
                                    else:
                                        duplicate_count += 1
                                else:
                                    filtered_records.append(record)
                                
                                progress = (i + 1) / total_records
                                progress_bar.progress(progress)
                                status_text.text(f"Processing {i+1} of {total_records} records...")
                            
                            status_placeholder.info(f"Found {len(filtered_records)} unique records to insert. Skipping {duplicate_count} duplicates.")
                            
                            if filtered_records:
                                batch_size = 100
                                success_count = 0
                                
                                for i in range(0, len(filtered_records), batch_size):
                                    batch = filtered_records[i:i+batch_size]
                                    
                                    if batch:
                                        response = supabase.table(TABLE_NAME).insert(batch).execute()
                                        
                                        if hasattr(response, 'data') and response.data:
                                            success_count += len(batch)
                                    
                                    progress = min(i + batch_size, len(filtered_records)) / max(1, len(filtered_records))
                                    progress_bar.progress(progress)
                                    status_text.text(f"Uploaded {success_count} of {len(filtered_records)} records...")
                                
                                st.toast(f"Field Result Updated! {success_count} unique records uploaded successfully.")
                            else:
                                st.warning("No new unique records to upload.")
                            
                            button_placeholder.button("Upload Complete!", disabled=True, key="complete_button")
                                
                        except Exception as e:
                            st.error(f"Error uploading field result: {str(e)}")
                            import traceback
                            st.code(traceback.format_exc())
                            
                            button_placeholder.button("Upload Failed - Try Again", key="retry_button")
                else:
                    st.error("Required columns not found in the uploaded file. Please ensure the file contains: chcode, status, SUB STATUS, DATE, and TIME columns.")
            except Exception as e:
                st.error(f"Error processing Excel file: {str(e)}")
        
        if upload_dataset:
            TABLE_NAME = 'rob_bike_dataset'
            try:
                xls = pd.ExcelFile(upload_dataset)
                df = pd.read_excel(xls)
                    
                df_clean = df.replace({np.nan: 0})
            
                df_filtered = df_clean.copy()
                
                st.subheader("Uploaded Dataset:")
                st.dataframe(df_filtered)
                
                possible_column_variants = {
                    'ChCode': ['ChCode'],
                    'Account Number': ['Account Number', 'Account_Number'],
                    'Client Name': ['Client Name', 'Client_Name'],
                    'Endorsement Date': ['Endorsement Date', 'Endorsement_Date'],
                    'Endrosement DPD': ['Endrosement DPD', 'Endrosement_DPD'],
                    'Store': ['Store'],
                    'Cluster': ['Cluster']
                }
                
                target_columns = [
                    'chcode',
                    'account_number',
                    'client_name',
                    'endo_date',
                    'endo_dpd',
                    'stores',
                    'cluster'
                ]
                
                column_mapping = {}
                for (key, variants), target in zip(possible_column_variants.items(), target_columns):
                    for variant in variants:
                        if variant in df_filtered.columns:
                            column_mapping[variant] = target
                            break 
                        
                if len(column_mapping) == len(target_columns):
                    df_selected = df_filtered[list(column_mapping.keys())].rename(columns=column_mapping)
                    
                    df_selected = df_selected.rename(columns=column_mapping)
                    
                    button_placeholder = st.empty()
                    status_placeholder = st.empty()
                    
                    upload_button = button_placeholder.button("Upload to Database", key="upload_dataset_button")
                    
                    if upload_button:
                        button_placeholder.button("Processing...", disabled=True, key="processing_dataset_button")
                        
                        try:
                            unique_id_col = 'account_number'
                            unique_ids = df_selected[unique_id_col].unique().tolist()
                            
                            for col in df_selected.columns:
                                if pd.api.types.is_datetime64_any_dtype(df_selected[col]):
                                    df_selected[col] = df_selected[col].dt.strftime('%Y-%m-%d')
                            
                            df_selected = df_selected.astype(object).where(pd.notnull(df_selected), None)
                            
                            new_records = df_selected.to_dict(orient="records")
                            
                            existing_records = []
                            batch_size_for_query = 20 
                            
                            progress_bar = st.progress(0)
                            status_text = status_placeholder.empty()
                            status_text.text("Fetching existing records...")
                            
                            for i in range(0, len(unique_ids), batch_size_for_query):
                                batch_ids = unique_ids[i:i+batch_size_for_query]
                                batch_ids = [id for id in batch_ids if id is not None and str(id).strip() != '']
                                
                                if batch_ids:
                                    try:
                                        batch_response = supabase.table(TABLE_NAME).select("*").in_(unique_id_col, batch_ids).execute()
                                        
                                        if hasattr(batch_response, 'data') and batch_response.data:
                                            existing_records.extend(batch_response.data)
                                    except Exception as e:
                                        st.warning(f"Error fetching batch {i}: {str(e)}. Continuing...")
                                
                                progress_value = min(1.0, (i + batch_size_for_query) / max(1, len(unique_ids)))
                                progress_bar.progress(progress_value)
                            
                            existing_df = pd.DataFrame(existing_records) if existing_records else pd.DataFrame()
                            
                            records_to_insert = []
                            records_to_update = []
                            total_records = len(new_records)
                            processed_count = 0
                            
                            status_text.text("Identifying records to insert or update...")
                            progress_bar.progress(0)
                            
                            def records_differ(new_record, existing_record):
                                for key, value in new_record.items():
                                    if key in existing_record and str(value) != str(existing_record[key]):
                                        return True
                                return False
                            
                            for new_record in new_records:
                                processed_count += 1
                                
                                if not existing_df.empty:
                                    matching_records = existing_df[existing_df[unique_id_col] == new_record[unique_id_col]]
                                    
                                    if not matching_records.empty:
                                        existing_record = matching_records.iloc[0].to_dict()
                                        if records_differ(new_record, existing_record):
                                            new_record['id'] = existing_record['id']
                                            records_to_update.append(new_record)
                                    else:
                                        records_to_insert.append(new_record)
                                else:
                                    records_to_insert.append(new_record)
                                    
                                progress_value = min(1.0, processed_count / total_records)
                                progress_bar.progress(progress_value)
                            
                            status_placeholder.info(f"Found {len(records_to_insert)} records to insert and {len(records_to_update)} records to update.")
                            
                            batch_size_for_db = 100
                            success_count = 0
                            
                            if records_to_insert:
                                status_text.text("Inserting new records...")
                                progress_bar.progress(0)
                                
                                for i in range(0, len(records_to_insert), batch_size_for_db):
                                    batch = records_to_insert[i:i+batch_size_for_db]
                                    
                                    if batch:
                                        try:
                                            response = supabase.table(TABLE_NAME).insert(batch).execute()
                                            
                                            if hasattr(response, 'data') and response.data:
                                                success_count += len(batch)
                                        except Exception as e:
                                            st.error(f"Error inserting records batch: {str(e)}")
                                    
                                    progress_value = min(1.0, min(i + batch_size_for_db, len(records_to_insert)) / max(1, len(records_to_insert)))
                                    progress_bar.progress(progress_value)
                                    status_text.text(f"Inserted {success_count} of {len(records_to_insert)} new records...")
                            
                            update_count = 0
                            if records_to_update:
                                status_text.text("Updating existing records...")
                                progress_bar.progress(0)
                                
                                for i, record in enumerate(records_to_update):
                                    record_id = record.pop('id') 
                                    
                                    try:
                                        response = supabase.table(TABLE_NAME).update(record).eq('id', record_id).execute()
                                        
                                        if hasattr(response, 'data') and response.data:
                                            update_count += 1
                                    except Exception as e:
                                        st.error(f"Error updating record {record_id}: {str(e)}")
                                    
                                    progress_value = min(1.0, (i + 1) / len(records_to_update))
                                    progress_bar.progress(progress_value)
                                    status_text.text(f"Updated {update_count} of {len(records_to_update)} existing records...")
                            
                            total_processed = success_count + update_count
                            if total_processed > 0:
                                st.toast(f"Dataset Updated! {success_count} records inserted and {update_count} records updated successfully.")
                                button_placeholder.button("Upload Complete!", disabled=True, key="complete_dataset_button")
                            else:
                                st.warning("No records were processed. Either no changes were needed or the operation failed.")
                                button_placeholder.button("Try Again", key="retry_dataset_button")
                                    
                        except Exception as e:
                            st.error(f"Error uploading dataset: {str(e)}")
                            import traceback
                            st.code(traceback.format_exc())
                            button_placeholder.button("Upload Failed - Try Again", key="error_dataset_button")
                else:
                    missing_cols = [col for col in source_columns if col not in df_filtered.columns]
                    st.error(f"Required columns not found in the uploaded file: {', '.join(missing_cols)}")
            except Exception as e:
                st.error(f"Error processing Excel file: {str(e)}")

    
    df = None
    sheet_names = []

    if uploaded_file is not None:
        file_content = uploaded_file.getvalue()
        file_buffer = io.BytesIO(file_content)
        
        try:
            xlsx = pd.ExcelFile(file_buffer)
            sheet_names = xlsx.sheet_names
            is_encrypted = False
            decrypted_file = file_buffer
            
        except Exception as e:
            if "file has been corrupted" in str(e) or "Workbook is encrypted" in str(e):
                is_encrypted = True
                st.sidebar.warning("This file appears to be password protected.")
                excel_password = st.sidebar.text_input("Enter Excel password", type="password")
                
                if not excel_password:
                    st.warning("Please enter the Excel file password.")
                    st.stop()
                
                try:
                    decrypted_file = io.BytesIO()
                    office_file = msoffcrypto.OfficeFile(io.BytesIO(file_content))
                    office_file.load_key(password=excel_password)
                    office_file.decrypt(decrypted_file)
                    decrypted_file.seek(0)
                    xlsx = pd.ExcelFile(decrypted_file)
                    sheet_names = xlsx.sheet_names
                except Exception as decrypt_error:
                    st.sidebar.error(f"Decryption failed: {str(decrypt_error)}")
                    st.stop()
            else:
                st.sidebar.error(f"Error reading file: {str(e)}")
                st.stop()
        
        selected_sheet = st.sidebar.selectbox(
            "Select Sheet", 
            options=sheet_names,
            index=0,
            key=f"{campaign}_sheet_selector"
        )
        
        try:
            if is_encrypted:
                decrypted_file.seek(0)
                df = pd.read_excel(decrypted_file, sheet_name=selected_sheet)
            else:
                df = pd.read_excel(xlsx, sheet_name=selected_sheet)
                
            if selected_sheet and preview:
                st.subheader(f"Preview of {selected_sheet}")
                df_preview = df.copy().dropna(how='all').dropna(how='all', axis=1)
                st.dataframe(df_preview, use_container_width=True)
                
        except Exception as e:
            st.sidebar.error(f"Error reading sheet: {str(e)}")
            
    with st.sidebar.expander("Data Cleaning Options"):
        remove_duplicates = st.checkbox("Remove Duplicates", value=False, key=f"{campaign}_remove_duplicates")
        remove_blanks = st.checkbox("Remove Blanks", value=False, key=f"{campaign}_remove_blanks")
        trim_spaces = st.checkbox("Trim Text", value=False, key=f"{campaign}_trim_spaces")
    
    with st.sidebar.expander("Data Manipulation"):
        st.markdown("#### Column Operations")
        enable_add_column = st.checkbox("Add Column", value=False)
        enable_column_removal = st.checkbox("Remove Column", value=False)
        enable_column_renaming = st.checkbox("Rename Column", value=False)
        
        st.markdown("#### Row Operations")
        enable_row_filtering = st.checkbox("Filter Row", value=False)
        enable_add_row = st.checkbox("Add Row", value=False)
        enable_row_removal = st.checkbox("Remove Row", value=False)
        
        st.markdown("#### Value Operations")
        enable_edit_values = st.checkbox("Edit Values", value=False)
    
    process_button = st.sidebar.button("Process File", type="primary", disabled=uploaded_file is None, key=f"{campaign}_process_button")

    if uploaded_file is not None:
        file_content = uploaded_file.getvalue() if hasattr(uploaded_file, 'getvalue') else uploaded_file.read()
        
        try:
            if "renamed_df" in st.session_state:
                df = st.session_state["renamed_df"]
            else:
                pass
            
            df = df.dropna(how='all', axis=0) 
            df = df.dropna(how='all', axis=1)

            if enable_add_column:
                st.subheader("Add New Columns")

                if "column_definitions" not in st.session_state:
                    st.session_state.column_definitions = []

                with st.form("add_column_form", clear_on_submit=True):
                    new_column_name = st.text_input("New Column Name")
                    column_source_type = st.radio("Column Source", ["Input Value", "Copy From Column", "Excel-like Formula"], key="source_type")

                    source_column = modification_type = prefix_text = suffix_text = selected_function = custom_function = formula = None
                    
                    if column_source_type == "Input Value":
                        input_value = st.text_input("Value to fill in each row")
                    elif column_source_type == "Copy From Column":
                        source_column = st.selectbox("Source Column (copy from)", df.columns.tolist(), key="source_column")
                        modification_type = st.radio("Modification Type", ["Direct Copy", "Text Prefix", "Text Suffix", "Apply Function"], key="mod_type")

                        if modification_type == "Text Prefix":
                            prefix_text = st.text_input("Prefix to add")
                        elif modification_type == "Text Suffix":
                            suffix_text = st.text_input("Suffix to add")
                        elif modification_type == "Apply Function":
                            function_options = ["To Uppercase", "To Lowercase", "Strip Spaces", "Custom Function"]
                            selected_function = st.selectbox("Select Function", function_options)
                            if selected_function == "Custom Function":
                                custom_function = st.text_area("Custom function (use 'x')", value="lambda x: x")
                    else:
                        st.info("Use column names in curly braces {} and expressions (e.g. `{Amount} * 2`, etc.)")
                        formula = st.text_area("Excel-like formula", height=80)

                    submitted = st.form_submit_button("Add to List")
                    if submitted and new_column_name:
                        st.session_state.column_definitions.append({
                            "name": new_column_name,
                            "source": column_source_type,
                            "source_column": source_column,
                            "modification_type": modification_type,
                            "prefix_text": prefix_text,
                            "suffix_text": suffix_text,
                            "function": selected_function,
                            "custom_function": custom_function,
                            "formula": formula,
                            "input_value": input_value if column_source_type == "Input Value" else None,
                        })
                        st.success(f"Queued column: {new_column_name}")

                if st.session_state.column_definitions:
                    st.write("🧾 Queued Columns to Add:")
                    for idx, col_def in enumerate(st.session_state.column_definitions):
                        st.markdown(f"- **{col_def['name']}** from **{col_def['source']}**")

                    if st.button("Apply All Column Additions"):
                            
                        try:
                            for col_def in st.session_state.column_definitions:
                                name = col_def["name"]
                                source = col_def["source"]
                                
                                if source == "Input Value":
                                    input_value = col_def["input_value"]
                                    df[name] = input_value
                                elif source == "Copy From Column":
                                    source_col = col_def["source_column"]
                                    mod_type = col_def["modification_type"]

                                    if mod_type == "Direct Copy":
                                        df[name] = df[source_col]
                                    elif mod_type == "Text Prefix":
                                        df[name] = col_def["prefix_text"] + df[source_col].astype(str)
                                    elif mod_type == "Text Suffix":
                                        df[name] = df[source_col].astype(str) + col_def["suffix_text"]
                                    elif mod_type == "Apply Function":
                                        if col_def["function"] == "To Uppercase":
                                            df[name] = df[source_col].astype(str).str.upper()
                                        elif col_def["function"] == "To Lowercase":
                                            df[name] = df[source_col].astype(str).str.lower()
                                        elif col_def["function"] == "Strip Spaces":
                                            df[name] = df[source_col].astype(str).str.strip()
                                        elif col_def["function"] == "Custom Function":
                                            func = eval(col_def["custom_function"])
                                            df[name] = df[source_col].apply(func)

                                elif source == "Excel-like Formula":
                                    formula = col_def["formula"]
                                    processed = formula
                                    for col in df.columns:
                                        pattern = r'\{' + re.escape(col) + r'\}'
                                        processed = re.sub(pattern, f"df['{col}']", processed)
                                    processed = processed.replace("IF(", "np.where(").replace("SUM(", "np.sum(")
                                    processed = processed.replace("AVG(", "np.mean(").replace("MAX(", "np.max(").replace("MIN(", "np.min(")
                                    df[name] = eval(processed)

                            st.success("All queued columns added successfully!")
                            st.session_state.renamed_df = df
                            st.session_state.column_definitions.clear()
                        except Exception as e:
                            st.error(f"Error applying column additions: {str(e)}")

            if enable_column_removal:
                st.subheader("Column Removal")
                cols = df.columns.tolist()
                cols_to_remove = st.multiselect("Select columns to remove", cols)
                if cols_to_remove:
                    df = df.drop(columns=cols_to_remove)
                    st.success(f"Removed columns: {', '.join(cols_to_remove)}")

            if enable_column_renaming:
                st.subheader("Column Renaming")
                
                rename_df = pd.DataFrame({
                    "original_name": df.columns,
                    "new_name": df.columns
                })
                
                edited_df = st.data_editor(
                    rename_df,
                    column_config={
                        "original_name": st.column_config.TextColumn("Original Column Name", disabled=True),
                        "new_name": st.column_config.TextColumn("New Column Name")
                    },
                    hide_index=True,
                    key="column_rename_editor"
                )
                
                if st.button("Apply Column Renames", key="apply_multiple_renames"):
                    rename_dict = {
                        orig: new 
                        for orig, new in zip(edited_df["original_name"], edited_df["new_name"]) 
                        if orig != new
                    }

                    if rename_dict:
                        df = df.rename(columns=rename_dict)
                        st.session_state["renamed_df"] = df
                        st.success(f"Renamed {len(rename_dict)} column(s): {', '.join([f'{k} → {v}' for k, v in rename_dict.items()])}")

            if enable_row_filtering:
                st.subheader("Row Filtering")
                filter_col = st.selectbox("Select column to filter by", df.columns.tolist())
                filter_value = st.text_input("Enter search/filter value")
                
                if filter_value and filter_col:
                    if pd.api.types.is_numeric_dtype(df[filter_col]):
                        try:
                            filter_value_num = float(filter_value)
                            filtered_df = df[df[filter_col] == filter_value_num]
                        except ValueError:
                            st.warning("Entered value is not numeric. Using string comparison instead.")
                            filtered_df = df[df[filter_col].astype(str).str.contains(filter_value, case=False, na=False)]
                    else:
                        filtered_df = df[df[filter_col].astype(str).str.contains(filter_value, case=False, na=False)]

                    st.write(f"Found {len(filtered_df)} rows matching filter: '{filter_value}' in column '{filter_col}'")
                    df = filtered_df
                    
            if enable_add_row:
                st.subheader("Add New Rows")
                with st.form("add_row_form"):
                    row_data = {}
                    for col in df.columns:
                        row_data[col] = st.text_input(f"Value for {col}", "")
                    
                    add_row_submitted = st.form_submit_button("Add Row")
                    
                    if add_row_submitted:
                        new_row = pd.DataFrame([row_data])
                        df = pd.concat([df, new_row], ignore_index=True)
                        st.success("Row added successfully!")
                        st.session_state["renamed_df"] = df

            if enable_row_removal:
                st.subheader("Remove Rows")
                st.info("Select rows to remove by index")
                
                with st.form("remove_row_form"):
                    row_indices = st.multiselect("Select row indices to remove", 
                                                options=list(range(len(df))),
                                                format_func=lambda x: f"Row {x}")
                    
                    remove_rows_submitted = st.form_submit_button("Remove Selected Rows")
                    
                    if remove_rows_submitted and row_indices:
                        df = df.drop(index=row_indices).reset_index(drop=True)
                        st.success(f"Removed {len(row_indices)} row(s)")
                        st.session_state["renamed_df"] = df

            if enable_edit_values:
                st.subheader("Edit Values")
                
                edited_df = st.data_editor(
                    df,
                    num_rows="dynamic",
                    use_container_width=True,
                    key="value_editor"
                )
                
                if st.button("Apply Value Changes"):
                    st.session_state["renamed_df"] = edited_df
                    st.success("Value changes applied!")
                    
            if enable_add_column or enable_column_removal or enable_column_renaming or enable_row_filtering or enable_add_row or enable_row_removal or enable_edit_values:
                buffer = io.BytesIO()
                df.to_excel(buffer, index=False, engine='openpyxl')
                file_content = buffer.getvalue()
                st.subheader("Modified Data Preview")
                st.dataframe(df.head(10), use_container_width=True)

        except Exception as e:
            st.error(f"Error loading or manipulating file: {str(e)}")

        if "renamed_df" in st.session_state:
            df = st.session_state["renamed_df"]
            buffer = io.BytesIO()
            df.to_excel(buffer, index=False, engine='openpyxl')
            buffer.seek(0)
            file_content = buffer.getvalue()

        if process_button and selected_sheet:
            try:
                with st.spinner("Processing file..."):
                    if automation_type == "Cured List":
                        result = processor.process_cured_list(
                            file_content, 
                            preview_only=False,
                            remove_duplicates=remove_duplicates, 
                            remove_blanks=remove_blanks, 
                            trim_spaces=trim_spaces
                        )
                        tabs = st.tabs(["Remarks", "Reshuffle", "Payments"])
                        with tabs[0]:
                            st.subheader("Remarks Data")
                            st.dataframe(result['remarks_df'], use_container_width=True)
                            st.download_button(label="Download Remarks File", data=result['remarks_binary'], file_name=result['remarks_filename'], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        with tabs[1]:
                            st.subheader("Reshuffle Data")
                            st.dataframe(result['others_df'], use_container_width=True)
                            st.download_button(label="Download Reshuffle File", data=result['others_binary'], file_name=result['others_filename'], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        with tabs[2]:
                            st.subheader("Payments Data")
                            st.dataframe(result['payments_df'], use_container_width=True)
                            st.download_button(label="Download Payments File", data=result['payments_binary'], file_name=result['payments_filename'], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        st.success("Cured List processed successfully!")
                    else:
                        if automation_type == "Data Clean":
                            result_df, output_binary, output_filename = getattr(processor, automation_map[automation_type])(
                                file_content,
                                preview_only=False,
                                remove_duplicates=remove_duplicates,
                                remove_blanks=remove_blanks,
                                trim_spaces=trim_spaces,
                                file_name=uploaded_file.name
                            )
                        elif automation_type == "Daily Remark Report":
                            result_df, output_binary, output_filename = getattr(processor, automation_map[automation_type])(
                                file_content, 
                                preview_only=False,
                                remove_duplicates=remove_duplicates, 
                                remove_blanks=remove_blanks, 
                                trim_spaces=trim_spaces,
                                report_date = report_date
                            )
                        else:
                            result_df, output_binary, output_filename = getattr(processor, automation_map[automation_type])(
                                file_content,
                                preview_only=False,
                                remove_duplicates=remove_duplicates,
                                remove_blanks=remove_blanks,
                                trim_spaces=trim_spaces
                            )
                        st.subheader("Processed Data")
                        st.dataframe(result_df, use_container_width=True)
                        st.download_button(label="Download File", data=output_binary, file_name=output_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        st.success(f"File processed successfully! Download '{output_filename}'")

                if "renamed_df" in st.session_state:
                    st.session_state.pop("renamed_df", None)

            except Exception as e:
                st.error(f"Error processing file: {str(e)}")

    st.sidebar.markdown("---")
    st.sidebar.markdown("© 2025 Automation Tool")

if __name__ == "__main__":
    main()
