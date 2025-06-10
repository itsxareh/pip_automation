import streamlit as st
import pandas as pd
import os
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, date, time
import io
import tempfile
import shutil
from processor.base import BaseProcessor

class PSBIAutoCuringProcessor(BaseProcessor):
    def process_new_endorsement(self, file_content, sheet_name=None, preview_only=False,
                         remove_duplicates=False, remove_blanks=False, trim_spaces=False, endo_date=None, preserve_colors=True):
        TABLE_NAME = 'psb_auto_dataset'
        all_account_numbers = []
        try:
            if isinstance(file_content, bytes):
                file_content = io.BytesIO(file_content)
            
            xls = pd.ExcelFile(file_content)
            
            df = pd.read_excel(
                xls, 
                sheet_name=sheet_name,
                dtype={'Account Number': str}
            )
            df = df.replace('', pd.NA)
            df = df.dropna(how='all')
            df = df.dropna(axis=1, how='all')
            df.columns = df.columns.str.strip()
            original_df = df.copy()
            
            df = self.clean_data(original_df, remove_duplicates, remove_blanks, trim_spaces)

            required_columns = [
                'Endorsement Date', 'Account Number', 'Endrosement OB',
                'Contact No.', 'BRAND', 'MODEL'        
            ]

            missing_cols = [col for col in required_columns if col not in df.columns]
            if missing_cols:
                st.error("Required columns not found in the uploaded file.")
                return None, None, None
            else:
                if 'Endorsement Date' in df.columns:
                    df = df.drop(columns='Endorsement Date')

                if 'Account Number 1' in df.columns:
                    df = df.drop(columns='Account Number 1')

                if 'Account Number' in df.columns:
                    df['Account Number'] = df['Account Number'].astype(str).str.strip()
                    account_numbers_list = [str(int(acc)) for acc in df['Account Number'].dropna().unique().tolist()]
                    
                    batch_size = 100 
                    existing_accounts = []
                    for i in range(0, len(account_numbers_list), batch_size):
                        batch = account_numbers_list[i:i + batch_size]
                        response = self.supabase.table('psb_auto_dataset').select('*').in_('account_number', batch).execute()

                        if hasattr(response, 'data') and response.data:
                            existing_accounts.extend([str(item['account_number']) for item in response.data])

                    initial_rows = len(df)
                    df = df[~df['Account Number'].astype(str).isin(existing_accounts)]
                    removed_rows = initial_rows - len(df)
                    
                    if removed_rows > 0:
                        st.write(f"Removed {removed_rows} rows with existing account numbers")
                    
                    if df.empty: 
                        st.warning("No new account numbers found (all account numbers exists)")
                        return None, None, None
                
                endo_date = endo_date.strftime('%m/%d/%Y')
                df.insert(0, 'ENDO DATE', endo_date)

                if 'Endrosement OB' in df.columns:
                    df['Endrosement OB'] = pd.to_numeric(df['Endrosement OB'], errors='coerce')
                    df.loc[df['Endrosement OB'].isna(), 'Endrosement OB'] = 1
                    df.loc[df['Endrosement OB'] == 0, 'Endrosement OB'] = 1

                    updated_ob_count = (df['Endrosement OB'] == 1).sum()
                    if updated_ob_count > 0:
                        st.warning(f"Updated {updated_ob_count} rows in 'Endrosement OB' (NaN or 0 → 1).")

                if 'Endrosement DPD' in df.columns:
                    df['Endrosement DPD'] = pd.to_numeric(df['Endrosement DPD'], errors='coerce')
                    df.loc[df['Endrosement DPD'].isna(), 'Endrosement DPD'] = 0

                    updated_dpd_count = (df['Endrosement DPD'] == 1).sum()
                    if updated_dpd_count > 0:
                        st.warning(f"Updated {updated_dpd_count} rows in 'Endrosement DPD' (NaN or 0 → 1).")
                
                if 'ENGINE NUMBER' in df.columns:
                    df['ENGINE NUMBER'] = df['ENGINE NUMBER'].apply(lambda x: np.nan if str(x).strip() == '0' else x)

                if 'CHASSIS NUMBER' in df.columns:
                    df['CHASSIS NUMBER'] = df['CHASSIS NUMBER'].apply(lambda x: np.nan if str(x).strip() == '0' else x)

                if preview_only:
                    return df, None, None
                
                bcrm_endo_df = df.copy()
                bcrm_endo_filename = f"rob_bike-new-({datetime.now().strftime('%Y-%m-%d')}).xlsx"
                
                cms_endo_df = df.copy()
                account_numbers = cms_endo_df['Account Number']
                all_account_numbers.extend(account_numbers.tolist())

                if 'Contact No.' in cms_endo_df.columns:
                    cms_endo_df['Contact No.'] = cms_endo_df['Contact No.'].apply(self.clean_phone_number)

                if 'BRAND' in cms_endo_df.columns and 'MODEL' in cms_endo_df.columns:
                    cms_endo_df['DESCRIP'] = cms_endo_df.apply(
                        lambda row: '' if (pd.isna(row['BRAND']) or pd.isna(row['MODEL']) or 
                                        str(row['BRAND']).strip() == '' or str(row['MODEL']).strip() == '' or
                                        str(row['BRAND']).lower() == 'nan' or str(row['MODEL']).lower() == 'nan')
                                    else f"{row['BRAND']} {row['MODEL']}", 
                        axis=1  
                    )
                    cols = cms_endo_df.columns.tolist()
                    cols.remove('DESCRIP')
                    cols.append('DESCRIP')
                    cms_endo_df = cms_endo_df[cols]

                def proper_case(name):
                    if pd.isna(name):
                        return ''
                    name = str(name).strip()
                    return ' '.join([part.capitalize() for part in name.split()])

                if 'ACCT NAME' in cms_endo_df.columns:
                    name_parts = cms_endo_df['ACCT NAME'].str.upper().str.split(', ', expand=True)

                    last_name = name_parts[0].apply(proper_case)
                    first_name = name_parts[1].apply(proper_case) if name_parts.shape[1] > 1 else ''
                    middle_name = name_parts[2].apply(proper_case) if name_parts.shape[1] > 2 else ''
                    
                    cms_endo_df.insert(3, 'FIRST NAME', first_name)
                    cms_endo_df.insert(4, 'MIDDLE NAME', middle_name)
                    cms_endo_df.insert(5, 'LAST NAME', last_name)

                unique_account_numbers = list(dict.fromkeys(all_account_numbers))
                if unique_account_numbers:
                    batch_size_for_query = 20
                    chcode_map = {}
                    
                    for i in range(0, len(unique_account_numbers), batch_size_for_query):
                        batch_ids = unique_account_numbers[i:i+batch_size_for_query]
                        batch_ids = [str(id).lstrip('0').strip() for id in batch_ids if id is not None and str(id).strip() != '']

                        if batch_ids:
                            try:
                                batch_response = self.supabase.table(TABLE_NAME).select("account_number, chcode").in_("account_number", batch_ids).execute()
                                if hasattr(batch_response, 'data') and batch_response.data:
                                    for record in batch_response.data:
                                        key = str(record['account_number']).strip()
                                        chcode_map[key] = str(record['chcode']).strip()
                            except Exception as e:
                                st.warning(f"Error fetching Ch Code batch {i}: {str(e)}. Continuing...")
                                
                    normalized_account_numbers = cms_endo_df['Account Number'].astype(str).str.lstrip('0').str.strip()
                    
                    cms_endo_df['CHCODE'] = normalized_account_numbers.map(chcode_map)

                    chcode_col = cms_endo_df.pop('CHCODE')
                    cms_endo_df.insert(2, 'CHCODE', chcode_col)

                cms_endo_df['AGENT FIRSTNAME'] = ''
                cms_endo_df['AGENT LASTNAME'] = ''

                cms_endo_filename = f"ROBBike-CMS-NewEndo-{datetime.now().strftime('%m-%d-%Y')}.xlsx"
                
                bcrm_endo_binary = self.create_excel_file(bcrm_endo_df)
                cms_endo_binary = self.create_excel_file(cms_endo_df)

                reshuffle_df = df[['Account Number', 'Endrosement OB']].copy()
                reshuffle_df = reshuffle_df.sort_values('Endrosement OB', ascending=True).reset_index(drop=True)
                reshuffle_df.rename(columns={'Account Number': 'Account No.'}, inplace=True)
                
                taggings = ['JDGANIAL', 'JAAGUILAR', 'NFMUANA']

                reshuffle_df['TAGGING'] = [taggings[i % len(taggings)] for i in range(len(reshuffle_df))]
                reshuffle_df = reshuffle_df.drop(columns=['Endrosement OB'])
                
                reshuffle_filename = f"ROBBike-CMS-Reshuffle-{datetime.now().strftime('%m-%d-%Y')}.xlsx"
                reshuffle_binary = self.create_excel_file(reshuffle_df)

                return {
                    'bcrm_endo_df': bcrm_endo_df,                
                    'bcrm_endo_binary': bcrm_endo_binary,                 
                    'bcrm_endo_filename': bcrm_endo_filename,
                    'cms_endo_df': cms_endo_df, 
                    'cms_endo_binary': cms_endo_binary,                 
                    'cms_endo_filename': cms_endo_filename,
                    'reshuffle_df': reshuffle_df, 
                    'reshuffle_binary': reshuffle_binary,                 
                    'reshuffle_filename': reshuffle_filename,
                }
            
        except Exception as e:
            st.error(f"Error processing new endorsement: {str(e)}")
            return None, None, None
        
    def clean_phone_number(self, phone):
        if pd.isna(phone) or str(phone).lower() == 'nan':
            return ''

        phone = str(phone)

        if '/' in phone:
            phone = phone.split('/')[0]

        digits = ''.join(c for c in phone if c.isdigit())

        match = re.search(r'9\d{9}$', digits)
        if match:
            return '0' + match.group() 
        else:
            return digits
        
    def create_excel_file(self, df):
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')

            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            final_columns = df.columns
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            account_col_idx = None
            maturity_col_idx = None
            endo_date_col_idx = None
            
            for i, col in enumerate(final_columns):
                col_letter = get_column_letter(i + 1)
                
                if col == 'Account Number':
                    account_col_idx = i + 1
                elif col == 'Maturity date':
                    maturity_col_idx = i + 1
                elif col == 'ENDO DATE':
                    endo_date_col_idx = i + 1
                
                if col in ['Account Number', 'ACCT NAME', 'Endrosement DPD', 'ENDO DATE', 'Endrosement OB', 
                          'MONTHLY AMORT', 'Maturity date', 'Contact No.', 'DESCRIP']:
                    max_length = max(
                        [len(str(cell.value)) if cell.value is not None else 0
                        for cell in worksheet[col_letter]]
                    )
                    adjusted_width = max_length + 2
                    worksheet.column_dimensions[col_letter].width = adjusted_width
                
                if col == 'Contact No.':
                    for row in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row, column=i+1)
                        cell.number_format = '@' 
                        if cell.value is not None:
                            cell.value = str(cell.value)

            for row in range(2, len(df) + 2):  
                if account_col_idx:
                    cell = worksheet.cell(row=row, column=account_col_idx)
                    cell.number_format = '@' 
                    if cell.value is not None:
                        cell.value = str(cell.value)
                        
                if maturity_col_idx:
                    cell = worksheet.cell(row=row, column=maturity_col_idx)
                    if cell.value is not None:
                        try:
                            date_value = pd.to_datetime(cell.value).strftime("%m/%d/%Y")
                            cell.value = date_value
                        except:
                            pass

                if endo_date_col_idx:
                    cell = worksheet.cell(row=row, column=endo_date_col_idx)
                    if cell.value is not None:
                        try:
                            date_value = pd.to_datetime(cell.value).strftime("%m/%d/%Y")
                            cell.value = date_value
                            cell.number_format = '@'
                        except:
                            pass

                for col_idx in range(1, len(final_columns) + 1):
                        cell = worksheet.cell(row=row, column=col_idx)
                        cell.border = thin_border

        output.seek(0)
        return output.getvalue()