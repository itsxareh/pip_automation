import streamlit as st
import pandas as pd
import os
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
import warnings
from datetime import datetime, date, time, timedelta
import io
import tempfile
import shutil
import re 
import msoffcrypto
import zipfile
from base import BaseProcessor

class BPIAutoCuringProcessor(BaseProcessor):
    
    def setup_directories(self, automation_type):
        """Create necessary directories based on automation type"""
        directories = {
            'updates': ["FOR_UPDATES", "BPI_FOR_UPDATES"],
            'uploads': ["FOR_UPLOADS", "BPI_FOR_UPLOADS"],
            'cured_list': ["CURED_LIST", "BPI_FOR_REMARKS", "BPI_FOR_PAYMENTS", "BPI_FOR_OTHERS"]
        }
        
        dirs_to_create = directories.get(automation_type, [])
        created_dirs = {}
        
        for dir_name in dirs_to_create:
            dir_path = os.path.join(self.temp_dir, dir_name)
            os.makedirs(dir_path, exist_ok=True)
            created_dirs[dir_name] = dir_path
            
        return created_dirs
        
    def process_updates_or_uploads(self, file_content, sheet_name, automation_type, preview_only=False,
                                   remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        try:
            byte_stream = io.BytesIO(file_content)
            xls = pd.ExcelFile(byte_stream)
            df = pd.read_excel(xls, sheet_name=sheet_name)
            
            required_columns = [
                'LAN', 'NAME', 'CTL4', 'PAST DUE', 'PAYOFF AMOUNT', 
                'PRINCIPAL', 'LPC', 'ADA SHORTAGE', 'UNIT', 'DPD',
                'EMAIL', 'CONTACT NUMBER 1', 'CONTACT NUMBER 2', 'ENDO DATE'
            ]
            
            df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)
            
            if preview_only:
                return df
                
            current_date = datetime.now().strftime('%m%d%Y')
            
            if automation_type == 'updates':
                output_filename = f"BPI AUTO CURING FOR UPDATES {current_date}.xlsx"
                input_filename = f"FOR UPDATE {current_date}.xlsx"
                dirs = self.setup_directories('updates')
                folder_key = 'BPI_FOR_UPDATES'
                input_folder_key = 'FOR_UPDATES'
            else: 
                output_filename = f"BPI AUTO CURING FOR UPLOADS {current_date}.xlsx"
                input_filename = f"FOR UPLOAD (NEW ENDO) {current_date}.xlsx"
                dirs = self.setup_directories('uploads')
                folder_key = 'BPI_FOR_UPLOADS'
                input_folder_key = 'FOR_UPLOADS'
            
            input_path = os.path.join(dirs[input_folder_key], input_filename)
            with open(input_path, 'wb') as f:
                f.write(file_content)
                
            column_map = {
                'EMAIL': 'EMAIL_ALS',
                'CONTACT NUMBER 1': 'MOBILE_NO_ALS',
                'CONTACT NUMBER 2': 'MOBILE_ALFES',
                'ENDO DATE': 'DATE REFERRED'
            }
            
            result_df = pd.DataFrame()
            
            for col in ['LAN', 'NAME', 'CTL4', 'PAST DUE', 'PAYOFF AMOUNT', 'PRINCIPAL', 'LPC', 
                        'ADA SHORTAGE', 'UNIT', 'DPD']:
                if col in df.columns:
                    result_df[col] = df[col].fillna("")
            
            result_df.insert(1, 'CH CODE', result_df['LAN'])
            
            for orig_col, new_col in column_map.items():
                if orig_col in df.columns:
                    if orig_col == 'CONTACT NUMBER 1' or orig_col == 'CONTACT NUMBER 2':
                        result_df[new_col] = df[orig_col].apply(lambda x: "" if pd.isna(x) else self.process_mobile_number(x))
                    elif orig_col == 'ENDO DATE':
                        result_df[new_col] = df[orig_col].apply(lambda x: self.format_date(x) if pd.notnull(x) else "")
                    else:
                        result_df[new_col] = df[orig_col].fillna("")
                else:
                    result_df[new_col] = ""
            
            result_df['LANDLINE_NO_ALFES'] = ""
            
            numeric_cols = ['PAST DUE', 'PAYOFF AMOUNT', 'PRINCIPAL', 'LPC', 'ADA SHORTAGE']
            for col in numeric_cols:
                if col in result_df.columns:
                    result_df[col] = pd.to_numeric(result_df[col], errors='coerce').fillna(0).round(2)
                    
            final_columns = [
                'LAN', 'CH CODE', 'NAME', 'CTL4', 'PAST DUE', 'PAYOFF AMOUNT', 'PRINCIPAL', 'LPC',
                'ADA SHORTAGE', 'EMAIL_ALS', 'MOBILE_NO_ALS', 'MOBILE_ALFES', 'LANDLINE_NO_ALFES', 
                'DATE REFERRED', 'UNIT', 'DPD'
            ]
            
            for col in final_columns:
                if col not in result_df.columns:
                    result_df[col] = ""
                    
            result_df = result_df[final_columns]
            
            output_path = os.path.join(dirs[folder_key], output_filename)
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Sheet1')
                
                worksheet = writer.sheets['Sheet1']
                for i, col in enumerate(final_columns):
                    max_length = max(
                        result_df[col].astype(str).map(len).max(),
                        len(col)
                    ) + 2
                    col_letter = chr(65 + i) 
                    worksheet.column_dimensions[col_letter].width = max_length
                    
                    if col in numeric_cols:
                        for row in range(2, len(result_df) + 2):
                            cell = worksheet[f"{col_letter}{row}"]
                            cell.number_format = '0.00'
                    
                    if col == 'DATE REFERRED':
                        for row in range(2, len(result_df) + 2):
                            cell = worksheet[f"{col_letter}{row}"]
                            value = cell.value
                            if value:
                                try: 
                                    cell.value = pd.to_datetime(value).strftime("%m/%d/%Y")
                                    cell.number_format = '@'
                                except:
                                    pass
            
            with open(output_path, 'rb') as f:
                output_binary = f.read()
                
            return result_df, output_binary, output_filename
            
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
            raise

    def process_updates(self, file_content, sheet_name=None, preview_only=False,
                        remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        return self.process_updates_or_uploads(file_content, sheet_name, 'updates', preview_only,
                                               remove_duplicates, remove_blanks, trim_spaces)
        
    def process_uploads(self, file_content, sheet_name=None, preview_only=False,
                        remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        return self.process_updates_or_uploads(file_content, sheet_name, 'uploads', preview_only,
                                               remove_duplicates, remove_blanks, trim_spaces)
    
    def process_cured_list(self, file_content, sheet_name=None, preview_only=False,
                           remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_input:
            temp_input.write(file_content)
            temp_input_path = temp_input.name
            
        try:
            xls = pd.ExcelFile(temp_input_path)
            df = pd.read_excel(xls, sheet_name=sheet_name)
            df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)
            
            if preview_only:
                return df
                
            current_date = datetime.now().strftime('%m%d%Y')
            dirs = self.setup_directories('cured_list')
            
            input_file = os.path.join(dirs["CURED_LIST"], f"CURED LIST {current_date}.xlsx")
            shutil.copy(temp_input_path, input_file)
            
            remarks_filename = f"BPI AUTOCURING REMARKS {current_date}.xlsx"
            others_filename = f"BPI AUTOCURING RESHUFFLE {current_date}.xlsx"
            payments_filename = f"BPI AUTOCURING PAYMENT {current_date}.xlsx"
            
            remarks_path = os.path.join(dirs["BPI_FOR_REMARKS"], remarks_filename)
            others_path = os.path.join(dirs["BPI_FOR_OTHERS"], others_filename)
            payments_path = os.path.join(dirs["BPI_FOR_PAYMENTS"], payments_filename)
            
            try:
                source_wb = openpyxl.load_workbook(input_file)
            except FileNotFoundError:
                print(f"Error: The file '{input_file}' was not found.")
                return
            
            dest_wb = openpyxl.Workbook()
            dest_ws = dest_wb.active
            
            headers = ["LAN", "Action Status", "Remark Date", "PTP Date", "Reason For Default", 
                    "Field Visit Date", "Remark", "Next Call Date", "PTP Amount", "Claim Paid Amount", 
                    "Remark By", "Phone No.", "Relation", "Claim Paid Date"]
            
            for col, header in enumerate(headers, 1):
                dest_ws.cell(row=1, column=col).value = header
            
            ws = source_wb.active
            if ws.max_column < 43:
                raise ValueError("File doesn't have the expected number of columns")
            
            current_row = 2
            total_rows = 0
            last_row = ws.max_row
            
            barcode_lookup = {}
            for row in range(2, last_row + 1):
                barcode = ws.cell(row=row, column=1).value
                phone1 = ws.cell(row=row, column=42).value
                phone2 = ws.cell(row=row, column=43).value
                phone1 = str(phone1).strip() if phone1 else ""
                phone2 = str(phone2).strip() if phone2 else ""
                if phone1:
                    phone1 = self.process_mobile_number(phone1)
                if phone2:
                    phone2 = self.process_mobile_number(phone2)
                if barcode:
                    barcode_lookup[barcode] = {
                        'date': ws.cell(row=row, column=3).value,
                        'amount': ws.cell(row=row, column=4).value,
                        'collector': ws.cell(row=row, column=2).value,
                        'phone1': phone1,
                        'phone2': phone2,
                    }
            
            nego_rows = []
            for row in range(2, last_row + 1):
                if (ws.cell(row=row, column=2).value != "SPMADRID" and 
                    (ws.cell(row=row, column=8).value is None or "PTP" not in str(ws.cell(row=row, column=8).value))):
                    nego_rows.append(row)
            
            if nego_rows:
                visible_count = len(nego_rows)
                
                for i, row_idx in enumerate(nego_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PTP NEW - CALL OUTS_PASTDUE"
                
                current_row += visible_count
                for i, row_idx in enumerate(nego_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PTP FF UP - CLIENT ANSWERED AND WILL SETTLE"
                
                current_row += visible_count
                
                for i, row_idx in enumerate(nego_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PAYMENT - CURED"
                
                current_row += visible_count
                
                total_rows += (visible_count * 3)
            
            ptp_rows = []
            for row in range(2, last_row + 1):
                if (ws.cell(row=row, column=2).value != "SPMADRID" and 
                    ws.cell(row=row, column=8).value is not None and 
                    "PTP" in str(ws.cell(row=row, column=8).value)):
                    ptp_rows.append(row)
            
            if ptp_rows:
                visible_count = len(ptp_rows)
                
                for i, row_idx in enumerate(ptp_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PTP FF UP - CLIENT ANSWERED AND WILL SETTLE"
                
                current_row += visible_count
                
                for i, row_idx in enumerate(ptp_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PAYMENT - CURED"
                
                current_row += visible_count
                
                total_rows += (visible_count * 2)
            
            spmadrid_rows = []
            for row in range(2, last_row + 1):
                if ws.cell(row=row, column=2).value == "SPMADRID":
                    spmadrid_rows.append(row)
            
            if spmadrid_rows:
                visible_count = len(spmadrid_rows)
                
                for i, row_idx in enumerate(spmadrid_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PTP NEW - CURED_GHOST"
                
                current_row += visible_count
                
                for i, row_idx in enumerate(spmadrid_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PAYMENT - CURED"
                
                current_row += visible_count
                
                total_rows += (visible_count * 2)
            
            final_row_count = total_rows + 1
            
            for row in range(2, final_row_count + 1):
                barcode = dest_ws.cell(row=row, column=1).value
                action_status = dest_ws.cell(row=row, column=2).value
                
                source_data = barcode_lookup.get(barcode, {})
                source_date = source_data.get('date')
                source_amount = source_data.get('amount')
                source_collector = source_data.get('collector')
                source_phone1 = source_data.get('phone1')
                source_phone2 = source_data.get('phone2')
                
                if source_date:
                    try:
                        if hasattr(source_date, 'strftime'): 
                            base_date = source_date
                        else:
                            try:
                                base_date = datetime.strptime(str(source_date), "%Y-%m-%d %H:%M:%S")
                            except:
                                try:
                                    base_date = datetime.strptime(str(source_date), "%Y-%m-%d")
                                except:
                                    base_date = datetime.now()
                    except:
                        base_date = datetime.now()
                    
                    if "PTP NEW" in action_status:
                        time_to_add = time(14, 40, 0)
                    elif "PTP FF" in action_status:
                        time_to_add = time(14, 50, 0)
                    elif "CURED" in action_status:
                        time_to_add = time(15, 0, 0)
                    else:
                        time_to_add = time(0, 0, 0)
                    
                    if not hasattr(base_date, 'time'):
                        base_date = datetime.combine(base_date, time(0, 0, 0))
                    
                    result_date = datetime.combine(base_date.date(), time_to_add)
                    
                    formatted_date = result_date.strftime("%m/%d/%Y %I:%M:%S %p")
                    dest_ws.cell(row=row, column=3).value = formatted_date
                    
                    formatted_date = result_date.strftime("%m/%d/%Y")
                    dest_ws.cell(row=row, column=4).value = formatted_date
                    
                    dest_ws.cell(row=row, column=3).number_format = '@'
                    dest_ws.cell(row=row, column=4).number_format = '@'
                else:
                    dest_ws.cell(row=row, column=3).value = ""
                    dest_ws.cell(row=row, column=4).value = ""
                
                phone_no = ""
                if "PAYMENT" not in action_status:
                    raw_value = dest_ws.cell(row=row, column=12).value
                    if raw_value:
                        phone_no = str(raw_value).strip().split('.')[0]
                
                if "PTP NEW" in action_status:
                    phone_value = source_phone1 if source_phone1 else source_phone2
                    remark_text = f"1_{phone_value} - PTP NEW"
                elif "PTP FF" in action_status:
                    phone_value = source_phone1 if source_phone1 else source_phone2
                    remark_text = f"{phone_value} - FPTP"
                elif "PAYMENT" in action_status:
                    remark_text = "CURED - CONFIRM VIA SELECTIVE LIST"
                else:
                    remark_text = ""
                
                dest_ws.cell(row=row, column=7).value = remark_text
                
                if "PAYMENT" in action_status:
                    dest_ws.cell(row=row, column=9).value = ""
                else:
                    dest_ws.cell(row=row, column=9).value = source_amount
                
                if "PAYMENT" in action_status:
                    dest_ws.cell(row=row, column=10).value = source_amount
                else:
                    dest_ws.cell(row=row, column=10).value = ""
                
                dest_ws.cell(row=row, column=11).value = source_collector
                
                if "PAYMENT" in action_status:
                    dest_ws.cell(row=row, column=12).value = ""
                else:
                    if source_phone1 and source_phone1 != "":
                        dest_ws.cell(row=row, column=12).value = source_phone1
                    else:
                        dest_ws.cell(row=row, column=12).value = source_phone2
                
                if "PAYMENT" in action_status and source_date:
                    if isinstance(source_date, datetime):
                        formatted_paid_date = source_date.strftime("%m/%d/%Y")
                    elif isinstance(source_date, date):
                        formatted_paid_date = source_date.strftime("%m/%d/%Y")
                    else:
                        try:
                            date_obj = datetime.strptime(str(source_date), "%Y-%m-%d %H:%M:%S")
                            formatted_paid_date = date_obj.strftime("%m/%d/%Y")
                        except:
                            try:
                                date_obj = datetime.strptime(str(source_date), "%Y-%m-%d")
                                formatted_paid_date = date_obj.strftime("%m/%d/%Y")
                            except:
                                formatted_paid_date = ""
                    dest_ws.cell(row=row, column=14).value = formatted_paid_date
                else:
                    dest_ws.cell(row=row, column=14).value = ""
            
            for row in range(2, final_row_count + 1):
                action_status = dest_ws.cell(row=row, column=2).value
                phone_no = dest_ws.cell(row=row, column=12).value
                
                if "PTP NEW" in action_status and phone_no:
                    dest_ws.cell(row=row, column=7).value = f"1_{phone_no} - PTP NEW"
                elif "PTP FF" in action_status and phone_no:
                    dest_ws.cell(row=row, column=7).value = f"{phone_no} - FPTP"
            
            for column in dest_ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                dest_ws.column_dimensions[column_letter].width = adjusted_width

            for row_idx in range(2, dest_ws.max_row + 1):  
                for col_idx in [3, 4, 14]: 
                    cell = dest_ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        cell_value_str = str(cell.value)
                        cell.value = cell_value_str
                        cell.number_format = '@'

            dest_wb.save(remarks_path)
            
            others_wb = openpyxl.Workbook()
            others_ws = others_wb.active
            
            others_ws.cell(row=1, column=1).value = ws.cell(row=1, column=1).value 
            others_ws.cell(row=1, column=2).value = "REMARK BY" 
            
            for row in range(2, last_row + 1):
                others_ws.cell(row=row, column=1).value = ws.cell(row=row, column=1).value

                reference_value = ws.cell(row=row, column=1).value 
                
                for cured_row in range(2, ws.max_row + 1):  
                    if ws.cell(row=cured_row, column=1).value == reference_value: 
                        others_ws.cell(row=row, column=2).value = ws.cell(row=cured_row, column=2).value 
                        break

            others_wb.save(others_path)
            
            payments_wb = openpyxl.Workbook()
            payments_ws = payments_wb.active
            payments_ws.cell(row=1, column=1).value = "LAN"
            payments_ws.cell(row=1, column=2).value = "ACCOUNT NUMBER"
            payments_ws.cell(row=1, column=3).value = "NAME"
            payments_ws.cell(row=1, column=4).value = "CARD NUMBER"
            payments_ws.cell(row=1, column=5).value = "PAYMENT AMOUNT"
            payments_ws.cell(row=1, column=6).value = "PAYMENT DATE"
            
            for row in range(2, last_row + 1):
                payments_ws.cell(row=row, column=1).value = ws.cell(row=row, column=17).value if ws.cell(row=row, column=17).value else ""
                payments_ws.cell(row=row, column=3).value = ws.cell(row=row, column=18).value if ws.cell(row=row, column=18).value else ""
                payments_ws.cell(row=row, column=5).value = ws.cell(row=row, column=4).value if ws.cell(row=row, column=4).value else ""
                date_value = ws.cell(row=row, column=3).value
                if date_value:
                    if isinstance(date_value, datetime):
                        formatted_date = date_value.strftime("%m/%d/%Y")
                    else:
                        formatted_date = str(date_value)
                    payments_ws.cell(row=row, column=6).value = formatted_date
            
            for row in range(2, last_row + 1):
                payments_ws.cell(row=row, column=6).number_format = "@"
            
            for column in payments_ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                payments_ws.column_dimensions[column_letter].width = adjusted_width
            
            payments_wb.save(payments_path)
            
            remarks_df = pd.read_excel(remarks_path)
            others_df = pd.read_excel(others_path)
            payments_df = pd.read_excel(payments_path)
            
            with open(remarks_path, 'rb') as f:
                remarks_binary = f.read()
            with open(others_path, 'rb') as f:
                others_binary = f.read()
            with open(payments_path, 'rb') as f:
                payments_binary = f.read()
                
            os.unlink(temp_input_path)
            
            return {
                'remarks_df': remarks_df, 
                'others_df': others_df, 
                'payments_df': payments_df,
                'remarks_binary': remarks_binary,
                'others_binary': others_binary,
                'payments_binary': payments_binary,
                'remarks_filename': remarks_filename,
                'others_filename': others_filename,
                'payments_filename': payments_filename
            }
        finally:
            if os.path.exists(temp_input_path):
                os.unlink(temp_input_path)
 