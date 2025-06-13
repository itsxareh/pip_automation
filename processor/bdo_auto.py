import streamlit as st
import pandas as pd
import os
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl import load_workbook
from datetime import datetime
import pytz
import io
import re 
import zipfile
from typing import List, Dict, Any, Optional
from processor.base import BaseProcessor as base
from supabase import create_client
from dotenv import load_dotenv
load_dotenv()


class BDOAutoProcessor(base):
    def get_previous_history(self, limit: int = 10) -> List[Dict[str, Any]]:
        try:
            response = (
                self.supabase
                .table("bdo_autoloan_inputset")
                .select(
                    "report_date, kept_count_b5, kept_bal_b5, alloc_bal_b5, "
                    "kept_count_b6, kept_bal_b6, alloc_bal_b6, created_at"
                )
                .order("created_at", desc=True)
                .limit(limit)
                .execute()
            )
            
            if response.data:
                return response.data
            else:
                return []
                
        except Exception as e:
            st.write(f"Error retrieving BDO Auto history: {str(e)}")
            return []
    
    def save_bdo_auto_data(self, kept_count_b5, kept_bal_b5, alloc_bal_b5, 
                          kept_count_b6, kept_bal_b6, alloc_bal_b6, report_date=None):
        try:
            if not report_date:
                report_date = datetime.now().strftime("%Y-%m-%d")
            else:
                try:
                    parsed_date = datetime.strptime(report_date, "%B %d")
                    parsed_date = parsed_date.replace(year=datetime.now().year)
                    report_date = parsed_date.strftime("%Y-%m-%d")
                except:
                    pass
            
            existing_response = (
                self.supabase
                .table("bdo_autoloan_inputset")
                .select("id")
                .eq("report_date", report_date)
                .execute()
            )
            
            data_payload = {
                "report_date": report_date,
                "kept_count_b5": kept_count_b5,
                "kept_bal_b5": kept_bal_b5,
                "alloc_bal_b5": alloc_bal_b5,
                "kept_count_b6": kept_count_b6,
                "kept_bal_b6": kept_bal_b6,
                "alloc_bal_b6": alloc_bal_b6,
                "updated_at": datetime.now().isoformat()
            }
            
            if existing_response.data and len(existing_response.data) > 0:
                existing_id = existing_response.data[0]["id"]
                response = (
                    self.supabase
                    .table("bdo_autoloan_inputset")
                    .update(data_payload)
                    .eq("id", existing_id)
                    .execute()
                )
                st.write(f"Updated existing BDO Auto data for date: {report_date}")
            else:
                data_payload["created_at"] = datetime.now().isoformat()
                response = (
                    self.supabase
                    .table("bdo_autoloan_inputset")
                    .insert(data_payload)
                    .execute()
                )
                st.write(f"Inserted new BDO Auto data for date: {report_date}")
            
            return True
                
        except Exception as e:
            st.write(f"Error saving BDO Auto data to Supabase: {str(e)}")
            return False
    
    def delete_bdo_auto_data(self, record_id: int) -> bool:
        try:
            response = (
                self.supabase
                .table("bdo_auto_loan_inputset")
                .delete()
                .eq("id", record_id)
                .execute()
            )
            return True
        except Exception as e:
            print(f"Error deleting BDO Auto data: {str(e)}")
            return False

    def process_agency_daily_report(self, file_content, sheet_name=None, preview_only=False,
        remove_duplicates=False, remove_blanks=False, trim_spaces=False, report_date=None,
        kept_count_b5=None, kept_bal_b5=None, alloc_bal_b5=None,
        kept_count_b6=None, kept_bal_b6=None, alloc_bal_b6=None):

        try:
            DIR = os.getcwd()
            
            TEMPLATE_DIR = os.path.join(DIR, "templates", "bdo_auto")
            daily_report_template = os.path.join(TEMPLATE_DIR, "AGENCY DAILY REPORT TEMPLATE.xlsx")
            daily_productivity_template = os.path.join(TEMPLATE_DIR, "DAILY PRODUCTIVITY TEMPLATE.xlsx")
            vs_report_template = os.path.join(TEMPLATE_DIR, "SPMADRID VS REPORT TEMPLATE.xlsx")
            
            if not os.path.exists(daily_report_template):
                st.error(f"Template file not found: {daily_report_template}")
                return None, None, None
                
            if not os.path.exists(daily_productivity_template):
                st.error(f"Template file not found: {daily_productivity_template}")
                return None, None, None
                
            try:
                with open(daily_report_template, 'rb') as template_file:
                    template_copy = io.BytesIO(template_file.read())
                try:
                    test_wb = load_workbook(template_copy)
                    test_wb.close()
                except zipfile.BadZipFile:
                    st.error(f"Template file is not a valid Excel file: {daily_report_template}")
                    return None, None, None
            except Exception as e:
                st.error(f"Error opening daily report template file: {str(e)}")
                return None, None, None
                
            try:
                with open(daily_productivity_template, 'rb') as template_file:
                    template_copy = io.BytesIO(template_file.read())
                try:
                    test_wb = load_workbook(template_copy)
                    test_wb.close()
                except zipfile.BadZipFile:
                    st.error(f"Template file is not a valid Excel file: {daily_productivity_template}")
                    return None, None, None
            except Exception as e:
                st.error(f"Error opening daily productivity template file: {str(e)}")
                return None, None, None
            
            BASE_DIR = os.path.join(DIR, "database", "bdo_auto")
            
            bucket_paths = {
                "Bucket 1": os.path.join(BASE_DIR, "BUCKET1_AGENT.xlsx"),
                "Bucket 2": os.path.join(BASE_DIR, "BUCKET2_AGENT.xlsx"),
                "Bucket 5&6": os.path.join(BASE_DIR, "BUCKET5&6_AGENT.xlsx")
            }
        
            bank_status_path = os.path.join(BASE_DIR, "BANK_STATUS.xlsx")
            rfd_list = os.path.join(BASE_DIR, "RFD_LISTS.xlsx")
            
            required_columns = [
                "Date", "Debtor", "Account No.", "Card No.", "Remark", "Remark By",
                "PTP Amount", "PTP Date", "Claim Paid Amount", "Claim Paid Date", 
                "Balance", "Status"
            ]
            
            bank_status_lookup = {}
            if os.path.exists(bank_status_path):
                df_bank_status = pd.read_excel(bank_status_path)
                if "CMS STATUS" not in df_bank_status.columns or "BANK STATUS" not in df_bank_status.columns:
                    st.error("Missing 'CMS STATUS' or 'BANK STATUS' column in BANK_STATUS.xlsx.")
                    return None, None, None
                bank_status_lookup = dict(zip(df_bank_status["CMS STATUS"].astype(str).str.strip(), 
                                            df_bank_status["BANK STATUS"].astype(str).str.strip()))
            else:
                st.error(f"Missing file: {bank_status_path}")
                return None, None, None
                
            rfd_valid_codes = set()
            if os.path.exists(rfd_list):
                df_rfd_list = pd.read_excel(rfd_list)
                if "RFD CODE" not in df_rfd_list.columns:
                    st.error("Missing 'RFD CODE' column in RFD_LISTS.xlsx.")
                    return None, None, None
                rfd_valid_codes = set(df_rfd_list["RFD CODE"].astype(str).str.upper())
            else:
                st.error(f"Missing file: {rfd_list}")
                return None, None, None
                
            byte_stream = io.BytesIO(file_content)
            xls = pd.ExcelFile(byte_stream)
            df_main = pd.read_excel(xls, sheet_name=sheet_name, dtype={"Account No.": str})
            
            df_main = self.clean_data(df_main, remove_duplicates, remove_blanks, trim_spaces)
            
            missing_columns = [col for col in required_columns if col not in df_main.columns]
            if missing_columns:
                st.error(f"Required columns not found in the uploaded file: {', '.join(missing_columns)}")
                return None, None, None
                
            df_main = df_main[required_columns]
            
            df_main["Remark By"] = df_main["Remark By"].astype(str).str.strip()
            
            df_main = df_main[~df_main["Remark"].isin([
                "Updates when case reassign to another collector", 
                "System Auto Update Remarks For PD"
            ])]
            
            df_main = df_main[~df_main["Card No."].isin([f"ch{i}" for i in range(1, 20)])]
            
            bucket_dfs = {}
            for bucket_name, bucket_path in bucket_paths.items():
                if os.path.exists(bucket_path):
                    df_bucket = pd.read_excel(bucket_path)
                    if "VOLARE USER" not in df_bucket.columns or "FULL NAME" not in df_bucket.columns:
                        st.warning(f"{bucket_name} missing required columns. Skipping.")
                        continue
                        
                    df_bucket["VOLARE USER"] = df_bucket["VOLARE USER"].astype(str).str.strip()
                    df_bucket["FULL NAME"] = df_bucket["FULL NAME"].astype(str).str.strip()
                    lookup_dict = dict(zip(df_bucket["VOLARE USER"], df_bucket["FULL NAME"]))
                    
                    matched_df = df_main[df_main["Remark By"].isin(df_bucket["VOLARE USER"])].copy()
                    matched_df["HANDLING OFFICER2"] = matched_df["Remark By"].map(lookup_dict)
                    
                    if bucket_name == "Bucket 1":
                        matched_df = matched_df[
                            (matched_df["Remark By"].isin(["SYSTEM", "LCMANZANO", "ACALVAREZ", "DSDEGUZMAN", "SRELIOT", "TANAZAIRE", "SPMADRID"]) &
                            matched_df["Card No."].astype(str).str.startswith("01")) |
                            (~matched_df["Remark By"].isin(["SYSTEM", "LCMANZANO", "ACALVAREZ", "DSDEGUZMAN", "SRELIOT", "TANAZAIRE", "SPMADRID"]))
                        ]
                    elif bucket_name == "Bucket 2":
                        matched_df = matched_df[
                            (matched_df["Remark By"].isin(["SYSTEM", "LCMANZANO", "ACALVAREZ", "DSDEGUZMAN", "SRELIOT", "TANAZAIRE", "SPMADRID"]) &
                            matched_df["Card No."].astype(str).str.startswith("02")) |
                            (~matched_df["Remark By"].isin(["SYSTEM", "LCMANZANO", "ACALVAREZ", "DSDEGUZMAN", "SRELIOT", "TANAZAIRE", "SPMADRID"]))
                        ]
                    elif bucket_name == "Bucket 5&6":
                        matched_df = matched_df[
                            (matched_df["Remark By"].isin(["SYSTEM", "LCMANZANO", "ACALVAREZ", "DSDEGUZMAN", "SRELIOT", "TANAZAIRE", "SPMADRID"]) &
                            matched_df["Card No."].astype(str).str.startswith(("05", "06"))) |
                            (~matched_df["Remark By"].isin(["SYSTEM", "LCMANZANO", "ACALVAREZ", "DSDEGUZMAN", "SRELIOT", "TANAZAIRE", "SPMADRID"]))
                        ]
                    
                    for col in ["PTP Date", "Claim Paid Date", "Date"]:
                        matched_df[col] = pd.to_datetime(matched_df[col], errors='coerce')
                    
                    matched_df["BANK STATUS"] = matched_df["Status"].astype(str).str.strip().map(bank_status_lookup)
                    
                    if not matched_df.empty:
                        bucket_dfs[bucket_name] = matched_df
                else:
                    st.error(f"Missing file: {bucket_path}")
            
            def extract_and_validate_rfd(remark):
                remark = str(remark).strip().rstrip("\\")
                rfd_match = re.search(r"RFD:\s*(\S+)$", remark)
                if rfd_match:
                    rfd = rfd_match.group(1).upper()
                else:
                    last_word = re.findall(r"\\\s*(\S+)", remark)
                    if last_word:
                        rfd = last_word[-1].upper()
                    else:
                        last_word = remark.split()[-1] if remark else np.nan
                        rfd = last_word.upper() if last_word else np.nan
                return rfd if rfd in rfd_valid_codes else np.nan
            
            def autofit_worksheet_columns(ws):
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    ws.column_dimensions[col_letter].width = adjusted_width
            
            def get_merged_cell_top_left(ws, cell_ref):
                """Find the top-left cell of a merged range containing the given cell_ref."""
                for merged_range in ws.merged_cells.ranges:
                    if cell_ref in merged_range:
                        return merged_range.min_row, merged_range.min_col
                return None, None  
            
            processed_dfs = {}
            for bucket_name, bucket_df in bucket_dfs.items():
                filtered_df = pd.DataFrame({
                    "Card Number": bucket_df["Card No."],
                    "PN": bucket_df["Account No."].astype(str).str.replace(r'\.0$', '', regex=True),
                    "NAME": bucket_df["Debtor"],
                    "BALANCE": bucket_df["Balance"].replace({',': ''}, regex=True).astype(float),
                    "HANDLING OFFICER2": bucket_df["HANDLING OFFICER2"].str.upper(),
                    "AGENCY3": "SP MADRID",
                    "STATUS4": bucket_df["BANK STATUS"],
                    "DATE OF CALL": bucket_df["Date"].dt.strftime("%m/%d/%Y"),
                    "PTP DATE": np.where(
                        bucket_df["PTP Date"].isna(),
                        np.where(bucket_df["Claim Paid Date"].isna(), np.nan, bucket_df["Claim Paid Date"].dt.strftime("%m/%d/%Y")),
                        bucket_df["PTP Date"].dt.strftime("%m/%d/%Y")
                    ),
                    "PTP AMOUNT": np.where(
                        bucket_df["PTP Amount"].isna() | (bucket_df["PTP Amount"] == 0),
                        np.where(bucket_df["Claim Paid Amount"].isna() | (bucket_df["Claim Paid Amount"] == 0), np.nan, bucket_df["Claim Paid Amount"]),
                        bucket_df["PTP Amount"]
                    ),
                    "RFD5": bucket_df["Remark"].apply(extract_and_validate_rfd)
                })
                
                filtered_df.reset_index(drop=True, inplace=True)
                
                for i in range(1, len(filtered_df)):
                    if filtered_df.loc[i, "HANDLING OFFICER2"] == "SYSTEM":
                        filtered_df.loc[i, "HANDLING OFFICER2"] = filtered_df.loc[i-1, "HANDLING OFFICER2"]
                
                filtered_df.loc[filtered_df["RFD5"].isna() & (filtered_df["STATUS4"] == "PTP"), "RFD5"] = "BUSY"
                filtered_df.loc[filtered_df["RFD5"].isna() & (filtered_df["STATUS4"] == "CALL NO PTP"), "RFD5"] = "NISV"
                filtered_df.loc[filtered_df["RFD5"].isna() & (filtered_df["STATUS4"] == "UNCON"), "RFD5"] = "NABZ"
                
                filtered_df.loc[(filtered_df["STATUS4"] == "PTP") & ((filtered_df["RFD5"] == "NISV") | (filtered_df["RFD5"] == "NABZ")), "RFD5"] = "BUSY"
                
                filtered_df['STATUS4'] = filtered_df['STATUS4'].replace('nan', np.nan)
                
                filtered_df = filtered_df[~(
                    filtered_df["STATUS4"].isna() | 
                    (filtered_df["STATUS4"] == "EXCLUDE")
                )]
                
                filtered_df.loc[filtered_df["STATUS4"] != "PTP", "PTP DATE"] = np.nan
                filtered_df.loc[filtered_df["STATUS4"] != "PTP", "PTP AMOUNT"] = np.nan
                
                ptp_complete_mask = (filtered_df["STATUS4"] == "PTP") & \
                    (filtered_df["PTP AMOUNT"].notna()) & \
                    (filtered_df["PTP DATE"].notna())

                if ptp_complete_mask.any():
                    ptp_complete_df = filtered_df[ptp_complete_mask].copy()

                    indices_to_drop = []
                    for pn in ptp_complete_df['PN'].unique():
                        pn_indices = ptp_complete_df[ptp_complete_df['PN'] == pn].index.tolist()
                        if len(pn_indices) > 1:
                            indices_to_drop.extend(pn_indices[:-1])

                    filtered_df = filtered_df.drop(indices_to_drop).reset_index(drop=True)
                
                processed_dfs[bucket_name] = filtered_df
            
            if preview_only:
                preview_data = {}
                for bucket_name, filtered_df in processed_dfs.items():
                    preview_df = filtered_df.drop(columns=["Card Number"])
                    preview_data[bucket_name] = preview_df.head(10)
                return preview_data, len(df_main), None

            bucket_5_6_df = processed_dfs.get("Bucket 5&6", pd.DataFrame())
            
            if not bucket_5_6_df.empty:
                bucket5_df = bucket_5_6_df[bucket_5_6_df["Card Number"].astype(str).str.startswith("05")].copy()
                bucket6_df = bucket_5_6_df[bucket_5_6_df["Card Number"].astype(str).str.startswith("06")].copy()
                
                bucket5_df = bucket5_df.drop(columns=["Card Number"])
                bucket6_df = bucket6_df.drop(columns=["Card Number"])
                
                if not report_date:
                    day = datetime.now().day
                    month = datetime.now().strftime("%B")
                    current_date = f"{month} {day}".upper()
                else:
                    current_date = report_date
                current_date_formatted = datetime.now().strftime("%m/%d/%Y") if not report_date else datetime.strptime(report_date, "%B %d").strftime("%m/%d/%Y")

                if current_date.endswith(" 0"):
                    current_date = current_date[:-2] + current_date[-1:]
                
                output_files = {}
                productivity_files = {}
                b5_prod_df = None
                b6_prod_df = None
                
                template_wb = load_workbook(daily_report_template)
                
                if not bucket5_df.empty:
                    wb5 = load_workbook(daily_report_template)
                    ws5 = wb5.active
                    
                    headers = bucket5_df.columns.tolist()
                    for col_idx, header in enumerate(headers, 1):
                        ws5.cell(row=1, column=col_idx, value=header)
                    
                    for r_idx, row in enumerate(bucket5_df.values, 2):
                        for c_idx, value in enumerate(row, 1):
                            ws5.cell(row=r_idx, column=c_idx, value=value)
                    
                    autofit_worksheet_columns(ws5)
                    
                    output_b5 = io.BytesIO()
                    wb5.save(output_b5)
                    output_b5.seek(0)
                    b5_binary = output_b5
                    output_files["B5"] = b5_binary.getvalue()
                    
                    wb5_prod = load_workbook(daily_productivity_template)
                    ws5_prod = wb5_prod.active
                    
                    row, col = get_merged_cell_top_left(ws5_prod, 'C2')
                    if row and col:
                        ws5_prod.cell(row=row, column=col, value=current_date_formatted)
                    else:
                        ws5_prod['C2'] = current_date_formatted
                    
                    ptp_rows_b5 = bucket5_df[bucket5_df["STATUS4"] == "PTP"]
                    ptp_count_b5 = len(ptp_rows_b5)
                    ptp_balance_sum_b5 = ptp_rows_b5["BALANCE"].sum() if ptp_count_b5 > 0 else 0.0
                    
                    b5_prod_df = pd.DataFrame({
                        "Date": [current_date_formatted],
                        "PTP Count": [ptp_count_b5],
                        "Balance Sum": [ptp_balance_sum_b5],
                        "Kept Count": [kept_count_b5],
                        "Kept Balance": [kept_bal_b5],
                        "Allocation Balance": [alloc_bal_b5]
                    })
                    
                    ws5_prod['F8'] = ptp_count_b5
                    ws5_prod['G8'] = ptp_balance_sum_b5
                    ws5_prod['G8'].number_format = "0.00"
                    ws5_prod["K8"] = kept_count_b5
                    ws5_prod["K9"] = kept_count_b5
                    ws5_prod["L8"] = kept_bal_b5
                    ws5_prod["C13"] = alloc_bal_b5

                    autofit_worksheet_columns(ws5_prod)
                    
                    output_b5_prod = io.BytesIO()
                    wb5_prod.save(output_b5_prod)
                    output_b5_prod.seek(0)
                    productivity_files["B5"] = output_b5_prod.getvalue()
                    
                if not bucket6_df.empty:
                    wb6 = load_workbook(daily_report_template)
                    ws6 = wb6.active
                    
                    headers = bucket6_df.columns.tolist()
                    for col_idx, header in enumerate(headers, 1):
                        ws6.cell(row=1, column=col_idx, value=header)
                    
                    for r_idx, row in enumerate(bucket6_df.values, 2):
                        for c_idx, value in enumerate(row, 1):
                            ws6.cell(row=r_idx, column=c_idx, value=value)
                    
                    autofit_worksheet_columns(ws6)
                    
                    output_b6 = io.BytesIO()
                    wb6.save(output_b6)
                    output_b6.seek(0)
                    b6_binary = output_b6
                    output_files["B6"] = b6_binary.getvalue()
                    
                    wb6_prod = load_workbook(daily_productivity_template)
                    ws6_prod = wb6_prod.active
                    
                    row, col = get_merged_cell_top_left(ws6_prod, 'C2')
                    if row and col:
                        ws6_prod.cell(row=row, column=col, value=current_date_formatted)
                    else:
                        ws6_prod['C2'] = current_date_formatted
                    
                    ptp_rows_b6 = bucket6_df[bucket6_df["STATUS4"] == "PTP"]
                    ptp_count_b6 = len(ptp_rows_b6)
                    ptp_balance_sum_b6 = ptp_rows_b6["BALANCE"].sum() if ptp_count_b6 > 0 else 0.0
                    
                    b6_prod_df = pd.DataFrame({
                        "Date": [current_date_formatted],
                        "PTP Count": [ptp_count_b6],
                        "Balance Sum": [ptp_balance_sum_b6],
                        "Kept Count": [kept_count_b6],
                        "Kept Balance": [kept_bal_b6],
                        "Allocation Balance": [alloc_bal_b6]
                    })
                    
                    ws6_prod['F8'] = ptp_count_b6
                    ws6_prod['G8'] = ptp_balance_sum_b6
                    ws6_prod['G8'].number_format = "0.00"
                    ws6_prod["K8"] = kept_count_b6
                    ws6_prod["K9"] = kept_count_b6
                    ws6_prod["L8"] = kept_bal_b6
                    ws6_prod["C13"] = alloc_bal_b6

                    autofit_worksheet_columns(ws6_prod)
                    
                    output_b6_prod = io.BytesIO()
                    wb6_prod.save(output_b6_prod)
                    output_b6_prod.seek(0)
                    productivity_files["B6"] = output_b6_prod.getvalue()
                                        
                    vs_report_wb = load_workbook(vs_report_template)
                    vs_report_ws = vs_report_wb.active

                    data = list(vs_report_ws.values)
                    if data:
                        headers = data[0]
                        rows = data[1:]
                        vs_df = pd.DataFrame(rows, columns=headers)
                    else:
                        vs_df = pd.DataFrame()

                    output_vs_report = io.BytesIO()
                    vs_report_wb.save(output_vs_report)
                    output_vs_report.seek(0)
                    vs_binary = output_vs_report
                    
                
                combined_output = io.BytesIO()
                with pd.ExcelWriter(combined_output, engine='openpyxl') as writer:
                    for bucket_name, filtered_df in processed_dfs.items():
                        output_df = filtered_df.drop(columns=["Card Number"])
                        output_df.to_excel(writer, index=False, sheet_name=bucket_name)
                combined_output.seek(0)
                
                temp_filename = f"temp_daily_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                b5_filename = f"AGENCY DAILY REPORT B5 AS OF {current_date}.xlsx"
                b6_filename = f"AGENCY DAILY REPORT B6 AS OF {current_date}.xlsx"
                b5_prod_filename = f"B5 Daily Productivity AS OF {current_date}.xlsx"
                b6_prod_filename = f"B6 Daily Productivity AS OF {current_date}.xlsx"
                vs_filename = f"SP MADRID VS AS OF {current_date}"

                if kept_count_b5 is not None and kept_count_b6 is not None:
                    save_success = self.save_bdo_auto_data(
                        kept_count_b5, kept_bal_b5, alloc_bal_b5,
                        kept_count_b6, kept_bal_b6, alloc_bal_b6,
                        report_date
                    )
                    if save_success:
                        st.write("Successfully saved BDO Auto data to Supabase")
                    else:
                        st.write("Failed to save BDO Auto data to Supabase")

                return {
                    "b5_df": bucket5_df,
                    "b6_df": bucket6_df,
                    "vs_df": vs_df,
                    "b5_prod_df": b5_prod_df,
                    "b6_prod_df": b6_prod_df,
                    "b5_binary": b5_binary.getvalue() if not bucket5_df.empty else None,
                    "b6_binary": b6_binary.getvalue() if not bucket6_df.empty else None,
                    "vs_binary": vs_binary.getvalue() if not vs_df.empty else None,
                    "b5_filename": b5_filename,
                    "b6_filename": b6_filename,
                    "vs_filename": vs_filename,
                    "b5_prod_binary": productivity_files.get("B5"),
                    "b6_prod_binary": productivity_files.get("B6"),
                    "b5_prod_filename": b5_prod_filename,
                    "b6_prod_filename": b6_prod_filename,
                    "preview": combined_output.getvalue(),
                    "temp_filename": temp_filename,
                    "output_files": output_files,
                    "productivity_files": productivity_files,
                    "output_filenames": {
                        "B5": b5_filename,
                        "B6": b6_filename,
                        "B5_Productivity": b5_prod_filename,
                        "B6_Productivity": b6_prod_filename
                    }
                }
                            
            return None, None, None
            
        except Exception as e:
            st.error(f"Error processing agency daily report: {str(e)}")
            return None, None, None

    def process_new_endorsement(self, file_content, sheet_name=None, preview_only=False,
                           remove_duplicates=False, remove_blanks=False, trim_spaces=False,
                           endo_date=None, bucket=None):
        TABLE_NAME = 'bdo_autoloan_dataset'
        all_account_numbers = []
        try:
            if isinstance(file_content, bytes):
                file_content = io.BytesIO(file_content)

            xls = pd.ExcelFile(file_content)

            df = pd.read_excel(
                xls,
                sheet_name=sheet_name,
                dtype={'PN': str, 'MOBILE NUMBER' : str}
            )
            df = df.replace('', pd.NA)
            df = df.dropna(how='all')
            df = df.dropna(axis=1, how='all')
            df.columns = df.columns.str.strip()
            original_df = df.copy()

            df = self.clean_data(original_df, remove_duplicates, remove_blanks, trim_spaces)

            required_columns = ['PN', 'COMPLETE_NAME', 'BALANCE', 'BUCKET', 'GROUP', "Due Date",
                            'MO_Amort', 'LAST_DATE', 'ZIP_CODE', "OVERDUE AMOUNT", 'ADDRESS',
                            'Email Address', 'MOBILE NUMBER', 'REMARKS']

            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                st.error(f"Required columns not found in the uploaded file: {', '.join(missing_columns)}")
                return None, None, None
            else:
                manila_timezone = pytz.timezone('Asia/Manila')
                current_datetime_manila = datetime.now(manila_timezone)
                current_date = current_datetime_manila.strftime('%m/%d/%Y')
                endo_date = endo_date.strftime('%m/%d/%Y') if endo_date else current_date

                bcrm_endo_columns = ['AGENCY', 'PN', 'BALANCE', 'PORT DATE', 'STATUS', 
                                    'PTC/PB', 'ENDO DATE', 'STATUS DATE', 'Delq. Reason', 'Customer Name',
                                    'Mobile', 'Home address', 'Office address', 'OD Amount2', 'EMAIL',
                                    'COBORROWER_EMAIL', 'OD_DAYS', 'COBORROWER_NAME', 'COBORROWER_NUMBER', 
                                    'LAST_BUCKET_AMOUNT', 'NEXT_PAYMENT_DUE_AMOUNT', 'EMPLOYER', 'AGENT', 'BUCKET']
                num_rows = len(df)
                bcrm_endo_df = pd.DataFrame(index=range(num_rows), columns=bcrm_endo_columns)

                bcrm_endo_df['AGENCY'] = 'SPM'
                bcrm_endo_df['STATUS'] = 'RETAINED'
                bcrm_endo_df['ENDO DATE'] = endo_date
                bcrm_endo_df['AGENT'] = 'PJIO'
                bcrm_endo_df['BUCKET'] = bucket

                if 'PN' in df.columns:
                    bcrm_endo_df['PN'] = df['PN'].astype(str)

                if 'BALANCE' in df.columns:
                    bcrm_endo_df['BALANCE'] = df['BALANCE']

                if 'COMPLETE_NAME' in df.columns:
                    bcrm_endo_df['Customer Name'] = df['COMPLETE_NAME']
                
                if 'MOBILE NUMBER' in df.columns:
                    bcrm_endo_df['Mobile'] = df['MOBILE NUMBER'].apply(self.process_mobile_number)

                if 'ADDRESS' in df.columns:
                    bcrm_endo_df['Home address'] = df['ADDRESS']

                if "OVERDUE AMOUNT" in df.columns:
                    bcrm_endo_df['OD Amount2'] = df["OVERDUE AMOUNT"]
                
                if 'Email Address' in df.columns:
                    bcrm_endo_df['EMAIL'] = df['Email Address']
                
                split_bucket = bucket[0] + bucket[-1] if bucket else 'NA'
                file_date_format = datetime.now().strftime('%Y-%m-%d')

                bcrm_endo_filename = f"bdo_auto_loan-new-({file_date_format}) {split_bucket}.xlsx"
                bcrm_endo_binary = self.create_excel_in_memory(bcrm_endo_df)
                
                cms_endo_columns = ['Bucket/Placement', 'Ch Code', 'Account Number', "Borrower's Name", 'Endo Date',
                                'Outstanding Balance', 'Overdue Balance', 'DPD', 'Monthly Amortization', 'Last Payment',
                                'Due Date', 'Contact Number', 'EMAIL', 'Address', 'Unit Model', 'Engine Number',
                                'Chassis Number', 'Plate Number']
                cms_endo_df = pd.DataFrame(index=range(num_rows), columns=cms_endo_columns)

                cms_endo_df['Bucket/Placement'] = bucket
                cms_endo_df['Endo Date'] = endo_date

                if 'PN' in df.columns:
                    cms_endo_df['Account Number'] = df['PN'].astype(str)
                elif 'ACCOUNT NUMBER' in df.columns:
                    cms_endo_df['Account Number'] = df['ACCOUNT NUMBER'].astype(str)
                else:
                    st.error("Missing 'PN' or 'ACCOUNT NUMBER' column in the uploaded file.")
                    return None, None, None

                account_numbers = cms_endo_df['Account Number'].astype(str).str.strip()
                all_account_numbers.extend(account_numbers.tolist())

                if 'COMPLETE_NAME' in df.columns:
                    cms_endo_df["Borrower's Name"] = df['COMPLETE_NAME']
                if 'BALANCE' in df.columns:
                    cms_endo_df['Outstanding Balance'] = df['BALANCE']
                if 'OVERDUE AMOUNT' in df.columns:
                    cms_endo_df['Overdue Balance'] = df['OVERDUE AMOUNT']
                if 'BUCKET' in df.columns:
                    cms_endo_df['DPD'] = df['BUCKET']
                if 'MO_Amort' in df.columns:
                    cms_endo_df['Monthly Amortization'] = df['MO_Amort']
                if 'LAST_DATE' in df.columns:
                    cms_endo_df['Last Payment'] = pd.to_datetime(df['LAST_DATE']).dt.strftime('%m/%d/%Y')
                if 'Due Date' in df.columns:
                    cms_endo_df['Due Date'] = pd.to_datetime(df['Due Date']).dt.strftime('%m/%d/%Y')
                if 'MOBILE NUMBER' in df.columns:
                    cms_endo_df['Contact Number'] = df['MOBILE NUMBER'].apply(self.process_mobile_number)
                if 'Email Address' in df.columns:
                    cms_endo_df['EMAIL'] = df['Email Address']
                if 'Model' in df.columns:
                    cms_endo_df['Unit Model'] = df['Model']
                if 'ADDRESS' in df.columns:
                    cms_endo_df['Address'] = df['ADDRESS']

                unique_account_numbers = list(dict.fromkeys(all_account_numbers))  
                if unique_account_numbers:
                    batch_size_for_query = 20
                    chcode_map = {}
                    
                    for i in range(0, len(unique_account_numbers), batch_size_for_query):
                        batch_ids = unique_account_numbers[i:i+batch_size_for_query]
                        batch_ids = [id for id in batch_ids if id is not None and id != '']
                        
                        if batch_ids:
                            try:
                                batch_response = self.supabase.table(TABLE_NAME).select("account_number, chcode").in_("account_number", batch_ids).execute()
                                if hasattr(batch_response, 'data') and batch_response.data:
                                    for record in batch_response.data:
                                        chcode_map[str(record['account_number']).strip()] = str(record['chcode']).strip()
                            except Exception as e:
                                st.warning(f"Error fetching Ch Code batch {i}: {str(e)}. Continuing...")
                    
                    cms_endo_df['Ch Code'] = cms_endo_df['Account Number'].apply(lambda x: chcode_map.get(str(x).strip(), ""))

                cms_endo_filename = f"BDO Auto {split_bucket} New Endo {file_date_format}.xlsx"
                cms_endo_binary = self.create_excel_in_memory(cms_endo_df)

                return {
                    'bcrm_endo_df': bcrm_endo_df,                
                    'bcrm_endo_binary': bcrm_endo_binary,                 
                    'bcrm_endo_filename': bcrm_endo_filename,
                    'cms_endo_df': cms_endo_df, 
                    'cms_endo_binary': cms_endo_binary,                 
                    'cms_endo_filename': cms_endo_filename,
                }

        except Exception as e:
            st.error(f"Error processing new endorsement: {str(e)}")
            return None, None, None

    def create_excel_in_memory(self, df):
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')

            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            final_columns = df.columns

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for i, col in enumerate(final_columns):
                if col in ['Due Date', 'Last Payment', 'ENDO DATE']:
                    col_letter = get_column_letter(i + 1)

                    for row in range(2, len(df) + 2):
                        cell = worksheet[f"{col_letter}{row}"]
                        if cell.value is not None:
                            try:
                                date_value = pd.to_datetime(cell.value).strftime("%m/%d/%Y")
                                cell.value = date_value
                                cell.number_format = '@'
                            except:
                                pass

        output.seek(0)
        return output.getvalue()

