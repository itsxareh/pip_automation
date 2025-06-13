import streamlit as st
import pandas as pd
import os
import io
from openpyxl import load_workbook
from processor.base import BaseProcessor

class SumishoProcessor(BaseProcessor):
    def process_daily_remark(self, file_content, sheet_name=None, preview_only=False,
    remove_duplicates=False, remove_blanks=False, trim_spaces=False,
    template_content=None, template_sheet=None, target_column=None):

        try:
            if isinstance(file_content, bytes):
                file_content = io.BytesIO(file_content)
                
            xls = pd.ExcelFile(file_content)
            df = pd.read_excel(xls, sheet_name=sheet_name)
            df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)

            if 'Date' not in df.columns or 'Remark' not in df.columns or 'Account No.' not in df.columns:
                raise ValueError("Required columns not found in the uploaded file.")
            
            df = df[df['Account No.'].notna()]                          
            df = df[df['Account No.'].astype(str).str.strip() != '']  

            if 'Time' in df.columns:
                if pd.api.types.is_object_dtype(df['Time']):
                    try:
                        df['Time'] = pd.to_datetime(df['Time'], format='%I:%M:%S %p')
                    except ValueError:
                        pass
                df = df.sort_values(by='Time', ascending=False)
                df = df.drop_duplicates(subset='Account No.', keep='first')
                
            df['FormattedDate'] = pd.to_datetime(df['Date']).dt.strftime('%m/%d/%Y')
            df['Date_Remark'] = df['FormattedDate'] + ' ' + df['Remark'].astype(str)
            
            account_remark_map = {}
            for idx, row in df.iterrows():
                account_number = str(int(row['Account No.']))
                formatted_date = row.get('FormattedDate')
                remark = row.get('Remark', '')

                if pd.isna(formatted_date):
                    value = str(remark) if pd.notna(remark) else ""
                else:
                    value = str(formatted_date) + ' ' + (str(remark) if pd.notna(remark) else "")
                    
                account_remark_map[account_number.strip()] = value
            
            if preview_only:
                template_stream = io.BytesIO(template_content)
                template_xls = pd.ExcelFile(template_stream)
                template_df = pd.read_excel(template_xls, sheet_name=template_sheet, header=1)
                
                account_number_col = None
                for col in template_df.columns:
                    col_str = str(col)
                    if 'ACCOUNT' in col_str.upper() and ('NUMBER' in col_str.upper() or 'NO' in col_str.upper()):
                        account_number_col = col
                        break
                        
                if not account_number_col:
                    st.write("Available columns:", template_df.columns.tolist())
                    raise ValueError("Account number column not found in template file.")
                    
                updated_count = 0
                for idx, row in template_df.iterrows():
                    template_acct = str(row[account_number_col]).strip() if pd.notna(row[account_number_col]) else ""
                    if template_acct in account_remark_map:
                        template_df.loc[idx, target_column] = account_remark_map[template_acct]
                        updated_count += 1
                st.write(f"Preview: {updated_count} cells would be updated in the template")
                return template_df
            
                date_report = pd.to_datetime(df['Date']).dt.strftime('%m%d%Y')
                
            output_filename = f"SP MADRID DAILY REPORT {date_report}1.xlsx"
            output_path = os.path.join(self.temp_dir, output_filename)
            
            template_stream = io.BytesIO(template_content)
            workbook = load_workbook(template_stream)
            
            if template_sheet in workbook.sheetnames:
                sheet = workbook[template_sheet]
            else:
                sheet = workbook.active
                st.warning(f"Sheet '{template_sheet}' not found, using active sheet instead")
            
            header_row = 2 
            account_col_idx = None
            target_col_idx = None
            
            for col_idx, cell in enumerate(sheet[header_row], 1):
                cell_value = str(cell.value).upper() if cell.value else ""
                if cell_value and ('ACCOUNT' in cell_value and ('NUMBER' in cell_value or 'NO' in cell_value)):
                    account_col_idx = col_idx
                if cell.value == target_column:
                    target_col_idx = col_idx

            if account_col_idx is None or target_col_idx is None:
                st.write("Header row content:", [cell.value for cell in sheet[header_row]])
                st.write(f"Looking for account column and target column: '{target_column}'")
                raise ValueError("Could not locate columns in Excel sheet")
                
            update_count = 0
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                account_cell = sheet.cell(row=row_idx, column=account_col_idx)
                
                if account_cell.value is not None:
                    account_str = str(account_cell.value).strip()
                    
                    if account_str in account_remark_map:
                        sheet.cell(row=row_idx, column=target_col_idx).value = account_remark_map[account_str]
                        update_count += 1

            st.info(f"Updated {update_count} cells in the Excel file")
            workbook.save(output_path)

            with open(output_path, 'rb') as f:
                output_binary = f.read()

            return None, output_binary, output_filename

        except Exception as e:
            st.error(f"Error processing daily report: {str(e)}")
            import traceback
            st.write(traceback.format_exc())
            raise