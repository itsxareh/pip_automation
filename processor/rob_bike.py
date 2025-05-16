import streamlit as st
import pandas as pd
import os
import numpy as np
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Border, Side, Alignment
from openpyxl.styles import numbers
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
from datetime import datetime
import io
from processor.base import BaseProcessor


from supabase import create_client
from dotenv import load_dotenv
load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)


class ROBBikeProcessor(BaseProcessor):
    def process_daily_remark(self, file_content, sheet_name=None, preview_only=False,
                    remove_duplicates=False, remove_blanks=False, trim_spaces=False, report_date=None):
        try:
            byte_stream = io.BytesIO(file_content)
            xls = pd.ExcelFile(byte_stream)
            df = pd.read_excel(xls, sheet_name=sheet_name)
            df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)
            
            required_columns = ['Time', 'Status', 'Account No.', 'Debtor', 'DPD', 'Remark', 'Remark By', 'PTP Amount', 'Balance', 'Claim Paid Amount']
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                st.error("Required columns not found in the uploaded file.")
                return None, None, None
            else: 
                if 'Time' in df.columns:
                    if pd.api.types.is_object_dtype(df['Time']):
                        try:
                            df['Time'] = pd.to_datetime(df['Time'], format='%I:%M:%S %p')
                        except ValueError:
                            pass
                    df = df.sort_values(by='Time', ascending=False)
                
                if 'Status' in df.columns:
                    df['Status'] = df['Status'].fillna('')
                
                    dnc_mask = df['Status'].str.contains('DNC', case=False)
                    blank_mask = df['Status'].str.strip() == ''
                
                    removed_dnc_count = dnc_mask.sum()
                    removed_blank_count = blank_mask.sum()
                
                    df = df[~(dnc_mask | blank_mask)]
                    
                    disposition = supabase.table('rob_bike_disposition').select("disposition").execute()
                
                    if disposition.data is None:
                        valid_dispo = []
                    else:
                        valid_dispo = [record['disposition'] for record in disposition.data]
                
                    not_in_valid_dispo = ~df['Status'].isin(valid_dispo)
                    removed_invalid_dispo_count = not_in_valid_dispo.sum()
                    df = df[~not_in_valid_dispo]
                    
                if 'Account No.' in df.columns and 'Status' in df.columns:
                    initial_duplicates = df.duplicated(subset=['Account No.', 'Status']).sum()
                    df['COMBINED_KEY'] = df['Account No.'].astype(str) + '_' + df['Status'].astype(str)
                    #remaining_duplicates = df.duplicated(subset=['COMBINED_KEY']).sum()
                    df = df.drop_duplicates(subset=['COMBINED_KEY'])
                    df = df.drop(columns=['COMBINED_KEY'])
                
                if 'Remark' in df.columns:
                    system_auto_update_remarks = df['Remark'].str.contains('System Auto Update Remarks For PD', case=False, na=False)
                    system_auto_update_remarks_count = system_auto_update_remarks.sum()
                    df = df[~system_auto_update_remarks]
                
                if 'Remark By' in df.columns:
                    jerivera_remarks = df['Remark By'].str.contains('JERIVERA', case=False, na=False)
                    system_remarks_count = jerivera_remarks.sum()
                    df = df[~jerivera_remarks]
                    system_remarks = df['Remark By'].str.contains('SYSTEM', case=False, na=False)
                    system_remarks_count = system_remarks.sum()
                    df = df[~system_remarks]
                    
                if 'PTP Amount' in df.columns and 'Balance' in df.columns and 'Claim Paid Amount' in df.columns:
                    df['PTP Amount'] = pd.to_numeric(df['PTP Amount'].replace({',': ''}, regex=True), errors='coerce')
                    df['Balance'] = pd.to_numeric(df['Balance'].replace({',': ''}, regex=True), errors='coerce')
                    df['Claim Paid Amount'] = pd.to_numeric(df['Claim Paid Amount'].replace({',': ''}, regex=True), errors='coerce')
                
                if 'PTP Amount' in df.columns and 'Status' in df.columns:
                    voluntary_surrender_rows = df[df['Status'] == 'PTP - VOLUNTARY SURRENDER']

                    invalid_amount_rows = voluntary_surrender_rows[
                        (voluntary_surrender_rows['PTP Amount'].isna()) |
                        (voluntary_surrender_rows['PTP Amount'] == 0)
                    ]

                    if not invalid_amount_rows.empty:
                        st.warning(f"Found {len(invalid_amount_rows)} row(s) with 'PTP - VOLUNTARY SURRENDER' but 0 or missing 'PTP Amount'.")
                        st.dataframe(invalid_amount_rows, use_container_width=True)
                        
                st.write(f"Removed: {removed_dnc_count} DNC, {removed_blank_count} blank status, {removed_invalid_dispo_count} invalid disposition, {system_auto_update_remarks_count} system auto update remarks, {system_remarks_count} system remarks, {initial_duplicates} duplicates.")

                if preview_only:
                    return df, None, None
                
                output_template = "DAILY MONITORING PTP, DEPO & REPO REPORT TEMPLATE.xlsx"
                sheet1 = "MONITORING"
                sheet2 = "PTP"
                sheet3 = "REPO"
                sheet4 = "DEPO"
                sheet5 = "EOD"
                
                monitoring_columns = ['Account Name', 'Account Number', 'Principal', 'EndoDate', 'Stores', 
                                    'Cluster', 'DaysPastDue', 'Field Status', 'Field Substatus', 
                                    'Status', 'subStatus', 'Notes', 'BarcodeDate', 'PTP Amount', 'PTP Date']
                monitoring_df = pd.DataFrame(columns=monitoring_columns)
                
                ptp_columns = ['Account Name', 'AccountNumber', 'Status', 'subStatus', 'Amount', 
                            'StartDate', 'Notes', 'ResultDate', 'EndoDate']
                ptp_df = pd.DataFrame(columns=ptp_columns)
                
                eod_df = pd.DataFrame()
                
                if 'Debtor' in df.columns:
                    monitoring_df['Account Name'] = df['Debtor'].str.upper()
                
                if 'Account No.' in df.columns:
                    monitoring_df['Account Number'] = df['Account No.']
                
                if 'Balance' in df.columns:
                    monitoring_df['Principal'] = df['Balance']
                
                if 'DPD' in df.columns:
                    monitoring_df['DaysPastDue'] = df['DPD']
                
                if 'Status' in df.columns:
                    status_parts = df['Status'].str.split('-', n=1)
                    monitoring_df['Status'] = status_parts.str[0].str.strip()
                    monitoring_df['subStatus'] = status_parts.str[1].str.strip().where(status_parts.str.len() > 1, "")
                
                if 'Remark' in df.columns:
                    monitoring_df['Notes'] = df['Remark']
                
                if 'Date' in df.columns:
                    monitoring_df['BarcodeDate'] = pd.to_datetime(df['Date']).dt.strftime('%m/%d/%Y')
                
                if 'PTP Amount' in df.columns and 'Claim Paid Amount' in df.columns:
                    ptp_amount = df['PTP Amount']
                    ptp_date = pd.to_datetime(df['PTP Date'], errors='coerce')
                    claim_paid_amount = df['Claim Paid Amount']
                    claim_paid_date = pd.to_datetime(df['Claim Paid Date'], errors='coerce')
                    
                    monitoring_df['PTP Amount'] = np.where(
                        ptp_amount.notna() & (ptp_amount != 0),
                        ptp_amount,
                        np.where(
                            claim_paid_amount.notna() & (claim_paid_amount != 0),
                            claim_paid_amount,
                            ''
                        )
                    )
                    
                    monitoring_df['PTP Date'] = np.where(
                        ptp_date.notna(),
                        ptp_date.dt.strftime('%m/%d/%Y'),
                        np.where(
                            claim_paid_date.notna(),
                            claim_paid_date.dt.strftime('%m/%d/%Y'),
                            ''
                        )
                    )
                                    
                if 'Account No.' in df.columns:
                    account_numbers = [str(int(acc)) for acc in df['Account No.'].dropna().unique().tolist()]
                    dataset_response = supabase.table('rob_bike_dataset').select('*').in_('account_number', account_numbers).execute()
                    
                    if hasattr(dataset_response, 'data') and dataset_response.data:
                        dataset_df = pd.DataFrame(dataset_response.data)
                        monitoring_df['Account Number'] = monitoring_df['Account Number'].apply(lambda x: str(int(float(x))) if pd.notnull(x) else '')
                        
                        account_data_map = {}
                        chcode_list = []
                        
                        for _, row in dataset_df.iterrows():
                            account_no = str(row['account_number']).strip()
                            chcode = row.get('chcode', '')
                            
                            if chcode:
                                chcode_list.append(chcode)
                                
                            account_data_map[account_no] = {
                                'ChCode': chcode,
                                'AccountNumber': "00" + account_no,
                                'EndoDate': row.get('endo_date', ''),
                                'Stores': row.get('stores', ''),
                                'Cluster': row.get('cluster', '')
                            }
                        
                        if chcode_list:
                            try:
                                field_results_response = supabase.table('rob_bike_field_result').select('*').in_('chcode', chcode_list).execute()
                                
                                if hasattr(field_results_response, 'data') and field_results_response.data:
                                    field_results_df = pd.DataFrame(field_results_response.data)
                                    
                                    if 'inserted_date' in field_results_df.columns:
                                        field_results_df['inserted_date'] = pd.to_datetime(field_results_df['inserted_date'])
                                    
                                    latest_status_map = {}
                                    
                                    if 'inserted_date' in field_results_df.columns:
                                        for chcode, group in field_results_df.groupby('chcode'):
                                            latest_row = group.sort_values('inserted_date', ascending=False).iloc[0]
                                            
                                            status = latest_row.get('status', '')
                                            substatus = latest_row.get('substatus', '')
                                            
                                            if status in ('0', '') or substatus in ('0', ''):
                                                status, substatus = '', ''
                                            
                                            latest_status_map[chcode] = {
                                                'Field_Status': status if status not in ('0', '') else '',
                                                'Field_Substatus': substatus if substatus not in ('0', '') else '',
                                            }
                                        
                                        for account_no, data in account_data_map.items():
                                            chcode = data['ChCode']
                                            if chcode in latest_status_map:
                                                account_data_map[account_no].update({
                                                    'Field_Status': latest_status_map[chcode]['Field_Status'],
                                                    'Field_Substatus': latest_status_map[chcode]['Field_Substatus'],
                                                })
                                            else:
                                                account_data_map[account_no].update({
                                                    'Field_Status': '',
                                                    'Field_Substatus': '',
                                                })
                                            
                            except Exception as e:
                                st.error(f"Error fetching field results: {str(e)}")
                        
                        monitoring_df['EndoDate'] = monitoring_df['Account Number'].map(
                            lambda acc_no: account_data_map.get(acc_no, {}).get('EndoDate', ''))
                        monitoring_df['EndoDate'] = pd.to_datetime(monitoring_df['EndoDate']).dt.strftime('%m/%d/%Y')
                        
                        monitoring_df['Stores'] = monitoring_df['Account Number'].map(
                            lambda acc_no: '' if account_data_map.get(acc_no, {}).get('Stores') in ['0', 0] 
                            else account_data_map.get(acc_no, {}).get('Stores', '')
                        )
                        
                        monitoring_df['Cluster'] = monitoring_df['Account Number'].map(
                            lambda acc_no: '' if account_data_map.get(acc_no, {}).get('Cluster') in ['0', 0] 
                            else account_data_map.get(acc_no, {}).get('Cluster', '')
                        )
                        
                        monitoring_df['Field Status'] = monitoring_df['Account Number'].map(
                            lambda acc_no: account_data_map.get(acc_no, {}).get('Field_Status', ''))
                        
                        monitoring_df['Field Substatus'] = monitoring_df['Account Number'].map(
                            lambda acc_no: account_data_map.get(acc_no, {}).get('Field_Substatus', ''))
                        
                        monitoring_df['Account Number'] = monitoring_df['Account Number'].map(
                            lambda acc_no: account_data_map.get(acc_no, {}).get('AccountNumber', ''))
                        
                ptp_data = df[df['Status'].str.contains('PTP', case=False, na=False)].copy() if 'Status' in df.columns else pd.DataFrame()
                
                if not ptp_data.empty:
                    if 'Debtor' in ptp_data.columns:
                        ptp_df['Account Name'] = ptp_data['Debtor'].str.upper()
                    
                    if 'Account No.' in ptp_data.columns:
                        ptp_df['AccountNumber'] = ptp_data['Account No.']
                    
                    if 'Status' in ptp_data.columns:
                        status_parts = ptp_data['Status'].str.split('-', n=1)
                        ptp_df['Status'] = status_parts.str[0].str.strip()
                        ptp_df['subStatus'] = status_parts.str[1].str.strip().where(status_parts.str.len() > 1, "")
                    
                    if 'PTP Amount' in ptp_data.columns:
                        ptp_df['Amount'] = ptp_data['PTP Amount']
                    
                    if 'PTP Date' in ptp_data.columns:
                        ptp_df['StartDate'] = pd.to_datetime(ptp_data['PTP Date']).dt.strftime('%Y-%m-%d')
                    
                    if 'Remark' in ptp_data.columns:
                        ptp_df['Notes'] = ptp_data['Remark']
                    
                    if 'Time' in ptp_data.columns:
                        time_only = pd.to_datetime(ptp_data['Time'], errors='coerce').dt.time

                        result_datetime = [
                            datetime.combine(report_date, t) if pd.notnull(t) else None for t in time_only
                        ]

                        ptp_df['ResultDate'] = [
                            dt.strftime('%m/%d/%Y %I:%M:%S %p').replace(' 0', ' ') if dt else '' for dt in result_datetime
                        ]
                        
                    if 'Account No.' in ptp_data.columns and 'account_data_map' in locals():
                        ptp_df['AccountNumber'] = ptp_df['AccountNumber'].apply(lambda x: str(int(float(x))) if pd.notnull(x) else '')
                        ptp_df['EndoDate'] = ptp_df['AccountNumber'].map(
                            lambda acc_no: account_data_map.get(acc_no, {}).get('EndoDate', ''))
                        ptp_df['EndoDate'] = pd.to_datetime(ptp_df['EndoDate']).dt.strftime('%m/%d/%Y')
                
                    if 'Account No.' in df.columns:
                        ptp_df['AccountNumber'] = ptp_df['AccountNumber'].map(
                            lambda acc_no: account_data_map.get(acc_no, {}).get('AccountNumber', ''))
            
                payment_statuses = [
                    "PAYMENT", "PAYMENT VIA CALL", "PAYMENT VIA SMS", "PAYMENT VIA EMAIL",
                    "PAYMENT VIA FIELD VISIT", "PAYMENT VIA CARAVAN", "PAYMENT VIA SOCMED"
                ]
                ptp_statuses = [
                    "PTP", "PTP VIA CALL", "PTP VIA SMS", "PTP VIA EMAIL", "PTP VIA FIELD VISIT",
                    "PTP VIA CARAVAN", "PTP VIA SOCMED"
                ]
                if 'Status' in df.columns:
                    status_parts = df['Status'].str.split('-', n=1)
                    df['Status'] = status_parts.str[0].str.strip()
                    
                    df['subStatus'] = status_parts.str[1].str.strip().where(status_parts.str.len() > 1, "")
                    
                df['Status'] = df['Status'].astype(str)
                df['subStatus'] = df['subStatus'].astype(str)
                
                total_principal = df['Balance'].sum()
                total_accounts = df['Balance'].count()
                
                filtered_vs = df[
                    (df['Status'].isin(payment_statuses)) &
                    (df['subStatus'].str.upper() == "VOLUNTARY SURRENDER")
                ]   
                repo_amount = filtered_vs['Balance'].sum()
                repo_count = filtered_vs['Balance'].count()

                filtered_payment = df[
                    (df['Status'].isin(ptp_statuses)) &
                    (~df['subStatus'].str.contains("Follow up", case=False, na=False))
                ]
                ptp_amount = filtered_payment['Balance'].sum()
                
                filtered_ptp = df[
                    (df['Status'].str.contains("PTP", case=False, na=False)) &
                    (~df['subStatus'].str.contains("Follow up", case=False, na=False))
                ]
                ptp_count = filtered_ptp.shape[0]
                
                eod_data = {
                    'Key': ['C2', 'D2', 'C5', 'D5', 'C9', 'D9'],
                    'Value': [total_principal, total_accounts, repo_amount, repo_count, ptp_amount, ptp_count]
                }
                eod_df = pd.DataFrame(eod_data)

                priority_substatus = [
                    ("FULLY PAID", "PAY OFF"),
                    ("PARTIAL", "STILL PD BUT WITH ARRANGEMENT"),
                    ("FULL UPDATE", "CURRENT")
                ]

                bottom_rows = []
                row_index = 12

                for substatus_value, label in priority_substatus:
                    temp_df = df[
                        (df['Status'].isin(payment_statuses)) &
                        (df['subStatus'].str.upper().str.contains(substatus_value.upper()))
                    ]
                    
                    for _, row in temp_df.iterrows():
                        bottom_rows.append({
                            'Key': f'C{row_index}',
                            'Value': row['Balance']
                        })
                        
                        ptp_value = row['Claim Paid Amount']
                        if ptp_value == 0 or ptp_value == '':
                            ptp_value = row['PTP Amount']
                            
                        bottom_rows.append({
                            'Key': f'D{row_index}',
                            'Value': ptp_value
                        })
                        
                        bottom_rows.append({
                            'Key': f'E{row_index}',
                            'Value': label
                        })
                        
                        row_index += 1

                min_rows = 2
                end_row = max(row_index, 12 + min_rows)
                for blank_row in range(row_index, end_row):
                    bottom_rows.append({'Key': f'C{blank_row}', 'Value': ''})
                    bottom_rows.append({'Key': f'E{blank_row}', 'Value': ''})
                    
                eod_df = pd.concat([eod_df, pd.DataFrame(bottom_rows)], ignore_index=True)
                
                template_path = os.path.join(os.getcwd(), "templates", "rob_bike", output_template)
                
                output_buffer = io.BytesIO()
                
                if os.path.exists(template_path):
                    try:
                        with open(template_path, 'rb') as template_file:
                            template_copy = io.BytesIO(template_file.read())
                            
                        try:
                            template_wb = load_workbook(template_copy)
                            
                            def append_df_to_sheet(sheet_name, df):
                                if sheet_name in template_wb.sheetnames:
                                    sheet = template_wb[sheet_name]
                                    start_row = sheet.max_row + 1
                                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row):
                                        for c_idx, value in enumerate(row, 1):
                                            sheet.cell(row=r_idx, column=c_idx).value = value
                            
                            append_df_to_sheet(sheet1, monitoring_df)
                            append_df_to_sheet(sheet2, ptp_df)
                            
                            def format_sheet(sheet_name, df=None):
                                sheet = template_wb[sheet_name]
                                
                                thin_border = Border(
                                    left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'),
                                )

                                if df is not None:
                                    for col_idx, col in enumerate(df.columns, 1):
                                        max_length = max(
                                            df[col].astype(str).map(len).max(),
                                            len(str(col))
                                        )
                                        adjusted_width = max_length + 2
                                        sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

                                    start_row = sheet.max_row - len(df) + 1
                                    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=len(df.columns)):
                                        for cell in row:
                                            cell.border = thin_border
                                            
                            if sheet5 in template_wb.sheetnames:
                                eod_sheet = template_wb[sheet5]
                                for _, row in eod_df.iterrows():
                                    cell_key = row['Key']
                                    value = row['Value']
                                    column_letter = cell_key[0]
                                    row_number = int(cell_key[1:])
                                    column_index = column_index_from_string(column_letter)
                                    eod_sheet.cell(row=row_number, column=column_index).value = value
                            
                            format_sheet(sheet1, monitoring_df)
                            format_sheet(sheet2, ptp_df)
                            format_sheet(sheet5, None)
                            
                            template_wb.save(output_buffer)
                            
                        except Exception as e:
                            st.error(f"Error processing template: {str(e)}")
                            
                    except Exception as e:
                        st.error(f"Error reading template file: {str(e)}")
                    
                else:
                    st.write("Template does not exist")
                    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                        monitoring_df.to_excel(writer, sheet_name=sheet1, index=False)
                        ptp_df.to_excel(writer, sheet_name=sheet2, index=False)
                        eod_df.to_excel(writer, sheet_name=sheet5, index=False)

                        workbook = writer.book
                        
                        workbook.create_sheet(title=sheet3)
                        workbook.create_sheet(title=sheet4)
                        
                        def format_sheet(sheet_name, df=None):
                            sheet = writer.sheets.get(sheet_name) or workbook[sheet_name]
                            
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'),
                            )
                            
                            if df is not None: 
                                for col_idx, col in enumerate(df.columns, 1):
                                    max_length = max(
                                        df[col].astype(str).map(len).max(),
                                        len(str(col))
                                    )
                                    adjusted_width = max_length + 2
                                    sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
                                
                                for row in sheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                                    for cell in row:
                                        cell.border = thin_border
                                        
                        eod_sheet = workbook[sheet5]
                        for _, row in eod_df.iterrows():
                            cell_key = row['Key']
                            value = row['Value']
                            column_letter = cell_key[0]
                            row_number = int(cell_key[1:])
                            column_index = column_index_from_string(column_letter)
                            eod_sheet.cell(row=row_number, column=column_index).value = value
                            
                        format_sheet(sheet1, monitoring_df)
                        format_sheet(sheet2, ptp_df)
                        format_sheet(sheet3)
                        format_sheet(sheet4)
                        format_sheet(sheet5, None)
                        
                output_buffer.seek(0)
                
                if not report_date:
                    report_date = datetime.now()

                date_str = report_date.strftime("%d%b%Y").upper()
                
                output_filename = f"DAILY MONITORING PTP, DEPO & REPO REPORT as of {date_str}.xlsx"
                
                return monitoring_df, output_buffer.getvalue(), output_filename
        
        except Exception as e:
            st.error(f"Error processing daily remark: {str(e)}")
            return None, None, None

    def process_new_endorsement(self, file_content, sheet_name=None, preview_only=False,
                         remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        try:
            if isinstance(file_content, bytes):
                file_content = io.BytesIO(file_content)
            
            xls = pd.ExcelFile(file_content)
            
            df = pd.read_excel(
                xls, 
                sheet_name=sheet_name,
                dtype={'Account Number': str}, 
                parse_dates=['Maturity date'] if sheet_name else None 
            )
            
            df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)
            
            if 'Endorsement Date' in df.columns:
                df = df.drop(columns='Endorsement Date')

            if 'Account Number 1' in df.columns:
                df = df.drop(columns='Account Number 1')

            if 'Account Number' in df.columns:
                df['Account Number'] = df['Account Number'].astype(str)
                account_numbers_list = df['Account Number'].dropna().unique().tolist()
                
                batch_size = 100 
                existing_accounts = []
                
                for i in range(0, len(account_numbers_list), batch_size):
                    batch = account_numbers_list[i:i + batch_size]
                    response = supabase.table('rob_bike_dataset').select('account_number').in_('account_number', batch).execute()
                    
                    if hasattr(response, 'data') and response.data:
                        st.write(response.data)
                        existing_accounts.extend([str(item['account_number']) for item in response.data])
                
                initial_rows = len(df)
                df = df[~df['Account Number'].astype(str).isin(existing_accounts)]
                removed_rows = initial_rows - len(df)
                
                if removed_rows > 0:
                    st.write(f"Removed {removed_rows} rows with existing account numbers")
                
                if df.empty:
                    st.warning("No new account numbers found (all account numbers exists)")
                    return None, None, None
            
            current_date = datetime.now().strftime('%Y/%m/%d')
            df.insert(0, 'ENDO DATE', current_date)
            
            if 'Endrosement OB' in df.columns:
                df['Endrosement OB'] = pd.to_numeric(df['Endrosement OB'], errors='coerce')
                zero_ob_rows = df[df['Endrosement OB'] == 0]
                if not zero_ob_rows.empty:
                    st.warning(f"Found {len(zero_ob_rows)} rows with 0 in Endorsement OB")
            
            if preview_only:
                return df, None, None
            
            result_df = df
            output_filename = f"rob_bike-new-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            output_path = os.path.join(os.getcwd(), output_filename)  
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Sheet1')

                workbook = writer.book
                worksheet = writer.sheets['Sheet1']
                final_columns = result_df.columns

                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                account_col_idx = None
                maturity_col_idx = None
                endo_date_col_idx = None
                
                for i, col in enumerate(final_columns):
                    col_letter = get_column_letter(i + 1)
                    
                    if col == 'Account Number':
                        account_col_idx = i + 1
                    elif col == 'Maturity date':
                        maturity_col_idx = i + 1
                    elif col == 'ENDO DATE':
                        endo_date_col_idx = i + 1
                    
                    if col in ['Account Number', 'ACCT NAME', 'Endrosement DPD', 'ENDO DATE', 'Endrosement OB', 'MONTHLY AMORT', 'Maturity date']:
                        max_length = max(
                            [len(str(cell.value)) if cell.value is not None else 0
                            for cell in worksheet[col_letter]]
                        )
                        adjusted_width = max_length + 2
                        worksheet.column_dimensions[col_letter].width = adjusted_width

                for row in range(2, len(result_df) + 2):  
                    if account_col_idx:
                        cell = worksheet.cell(row=row, column=account_col_idx)
                        cell.number_format = '@' 
                        if cell.value is not None:
                            cell.value = str(cell.value)
                            
                    if maturity_col_idx:
                        cell = worksheet.cell(row=row, column=maturity_col_idx)
                        if cell.value is not None:
                            try:
                                date_value = pd.to_datetime(cell.value).strftime("%m/%d/%Y")
                                cell.value = date_value
                            except:
                                pass

                    if endo_date_col_idx:
                        cell = worksheet.cell(row=row, column=endo_date_col_idx)
                        if cell.value is not None:
                            try:
                                date_value = pd.to_datetime(cell.value).strftime("%m/%d/%Y")
                                cell.value = date_value
                                cell.number_format = '@'
                            except:
                                pass
                    
                    for col_idx in range(1, len(final_columns) + 1):
                        cell = worksheet.cell(row=row, column=col_idx)
                        cell.border = thin_border

            with open(output_path, 'rb') as f:
                output_binary = f.read()
            
            return result_df, output_binary, output_filename
        
        except Exception as e:
            st.error(f"Error processing new endorsement: {str(e)}")
            return None, None, None