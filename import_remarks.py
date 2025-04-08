import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, date, time
import os

def import_data():
    """
    Python implementation to process Excel data.
    Takes a CURED LIST file and produces three output files:
    - BPI AUTOCURING REMARKS [date].xlsx
    - BPI AUTOCURING RESHUFFLE [date].xlsx
    - BPI AUTOCURING PAYMENT [date].xlsx
    """
    current_date = date.today().strftime("%m%d%Y")
    pwd = os.getcwd()
    pwd = os.path.normpath(pwd)

    cured_list_file = os.path.join(pwd,  "CURED_LIST", f"CURED LIST {current_date}.xlsx")
    remarks_folder = os.path.join(pwd,  "BPI_FOR_REMARKS")
    payments_folder = os.path.join(pwd,  "BPI_FOR_PAYMENTS") 
    others_folder = os.path.join(pwd, "BPI_FOR_OTHERS")

    for folder in [remarks_folder, payments_folder, others_folder]:
        if not os.path.exists(folder):
            os.makedirs(folder)

    remarks_filename = f"BPI AUTOCURING REMARKS {current_date}.xlsx"
    others_filename = f"BPI AUTOCURING RESHUFFLE {current_date}.xlsx"
    payments_filename = f"BPI AUTOCURING PAYMENT {current_date}.xlsx"

    remarks_path = os.path.join(remarks_folder, remarks_filename)
    payments_path = os.path.join(payments_folder, payments_filename)
    others_path = os.path.join(others_folder, others_filename)
    
    try:
        source_wb = openpyxl.load_workbook(cured_list_file)
    except FileNotFoundError:
        print(f"Error: The file '{cured_list_file}' was not found.")
        return
    
    dest_wb = openpyxl.Workbook()
    dest_ws = dest_wb.active
    
    headers = ["LAN", "Action Status", "Remark Date", "PTP Date", "Reason For Default", 
               "Field Visit Date", "Remark", "Next Call Date", "PTP Amount", "Claim Paid Amount", 
               "Remark By", "Phone No.", "Relation", "Claim Paid Date"]
    
    for col, header in enumerate(headers, 1):
        dest_ws.cell(row=1, column=col).value = header
    
    ws = source_wb.active
    
    current_row = 2
    total_rows = 0
    last_row = ws.max_row
    
    barcode_lookup = {}
    for row in range(2, last_row + 1):
        barcode = ws.cell(row=row, column=1).value
        if barcode:
            barcode_lookup[barcode] = {
                'date': ws.cell(row=row, column=3).value,  # Column C
                'amount': ws.cell(row=row, column=4).value,  # Column D
                'collector': ws.cell(row=row, column=2).value,  # Column B
                'phone1': ws.cell(row=row, column=42).value,  # Column AP
                'phone2': ws.cell(row=row, column=43).value,  # Column AQ
            }
            
    print("Processing NON-SPMADRID data...")
    nego_rows = []
    for row in range(2, last_row + 1):
        if (ws.cell(row=row, column=2).value != "SPMADRID" and 
            (ws.cell(row=row, column=8).value is None or "PTP" not in str(ws.cell(row=row, column=8).value))):
            nego_rows.append(row)
    
    if nego_rows:
        visible_count = len(nego_rows)
        
        for i, row_idx in enumerate(nego_rows):
            barcode = ws.cell(row=row_idx, column=1).value
            dest_ws.cell(row=current_row + i, column=1).value = barcode
            dest_ws.cell(row=current_row + i, column=2).value = "PTP NEW - CALL OUTS_PASTDUE"
        
        current_row += visible_count
        for i, row_idx in enumerate(nego_rows):
            barcode = ws.cell(row=row_idx, column=1).value
            dest_ws.cell(row=current_row + i, column=1).value = barcode
            dest_ws.cell(row=current_row + i, column=2).value = "PTP FF UP - CLIENT ANSWERED AND WILL SETTLE"
        
        current_row += visible_count
        
        for i, row_idx in enumerate(nego_rows):
            barcode = ws.cell(row=row_idx, column=1).value
            dest_ws.cell(row=current_row + i, column=1).value = barcode
            dest_ws.cell(row=current_row + i, column=2).value = "PAYMENT - CURED"
        
        current_row += visible_count
        
        total_rows += (visible_count * 3)
        print(f"NON-SPMADRID data processed: {visible_count * 3} rows.")
    else:
        print("No NON-SPMADRID found.")
    
    print("Processing PTP data...")
    ptp_rows = []
    for row in range(2, last_row + 1):
        if (ws.cell(row=row, column=2).value != "SPMADRID" and 
            ws.cell(row=row, column=8).value is not None and 
            "PTP" in str(ws.cell(row=row, column=8).value)):
            ptp_rows.append(row)
    
    if ptp_rows:
        visible_count = len(ptp_rows)
        
        for i, row_idx in enumerate(ptp_rows):
            barcode = ws.cell(row=row_idx, column=1).value
            dest_ws.cell(row=current_row + i, column=1).value = barcode
            dest_ws.cell(row=current_row + i, column=2).value = "PTP FF UP - CLIENT ANSWERED AND WILL SETTLE"
        
        current_row += visible_count
        
        for i, row_idx in enumerate(ptp_rows):
            barcode = ws.cell(row=row_idx, column=1).value
            dest_ws.cell(row=current_row + i, column=1).value = barcode
            dest_ws.cell(row=current_row + i, column=2).value = "PAYMENT - CURED"
        
        current_row += visible_count
        
        total_rows += (visible_count * 2)
        print(f"NON-SPMADRID PTP processed: {visible_count * 2} rows.")
    else:
        print("No NON-SPMADRID PTP found.")
    
    print("Processing SPMADRID data...")
    spmadrid_rows = []
    for row in range(2, last_row + 1):
        if ws.cell(row=row, column=2).value == "SPMADRID":
            spmadrid_rows.append(row)
    
    if spmadrid_rows:
        visible_count = len(spmadrid_rows)
        
        for i, row_idx in enumerate(spmadrid_rows):
            barcode = ws.cell(row=row_idx, column=1).value
            dest_ws.cell(row=current_row + i, column=1).value = barcode
            dest_ws.cell(row=current_row + i, column=2).value = "PTP NEW - CURED_GHOST"
        
        current_row += visible_count
        
        for i, row_idx in enumerate(spmadrid_rows):
            barcode = ws.cell(row=row_idx, column=1).value
            dest_ws.cell(row=current_row + i, column=1).value = barcode
            dest_ws.cell(row=current_row + i, column=2).value = "PAYMENT - CURED"
        
        current_row += visible_count
        
        total_rows += (visible_count * 2)
        print(f"SPMADRID data processed: {visible_count * 2} rows.")
    else:
        print("No SPMADRID data found.")
    
    print(f"Remarks complete. Total rows: {total_rows}")
    
    final_row_count = total_rows + 1
    
    for row in range(2, final_row_count + 1):
        barcode = dest_ws.cell(row=row, column=1).value
        action_status = dest_ws.cell(row=row, column=2).value
        
        source_data = barcode_lookup.get(barcode, {})
        source_date = source_data.get('date')
        source_amount = source_data.get('amount')
        source_collector = source_data.get('collector')
        source_phone1 = source_data.get('phone1')
        source_phone2 = source_data.get('phone2')

        if source_date:
            try:
                if hasattr(source_date, 'strftime'): 
                    base_date = source_date
                else:
                    try:
                        base_date = datetime.strptime(str(source_date), "%Y-%m-%d %H:%M:%S")
                    except:
                        try:
                            base_date = datetime.strptime(str(source_date), "%Y-%m-%d")
                        except:
                            base_date = datetime.now()
            except:
                base_date = datetime.now()
            
            if "PTP NEW" in action_status:
                time_to_add = time(14, 40, 0)
            elif "PTP FF" in action_status:
                time_to_add = time(14, 50, 0)
            elif "CURED" in action_status:
                time_to_add = time(15, 0, 0)
            else:
                time_to_add = time(0, 0, 0)
            
            if not hasattr(base_date, 'time'):
                base_date = datetime.combine(base_date, time(0, 0, 0))
            
            result_date = datetime.combine(base_date.date(), time_to_add)
            
            formatted_date = result_date.strftime("%m/%d/%Y %I:%M:%S %p")
            dest_ws.cell(row=row, column=3).value = formatted_date
            
            formatted_date = result_date.strftime("%m/%d/%Y")
            dest_ws.cell(row=row, column=4).value = formatted_date
            
            dest_ws.cell(row=row, column=3).number_format = '@'
            dest_ws.cell(row=row, column=4).number_format = '@'
        else:
            dest_ws.cell(row=row, column=3).value = ""
            dest_ws.cell(row=row, column=4).value = ""
        
        phone_no = ""
        if "PAYMENT" not in action_status:
            phone_no = dest_ws.cell(row=row, column=12).value
        
        if "PTP NEW" in action_status:
            phone_value = source_phone1 if source_phone1 else source_phone2
            remark_text = f"1_{phone_value} - PTP NEW"
        elif "PTP FF" in action_status:
            phone_value = source_phone1 if source_phone1 else source_phone2
            remark_text = f"{phone_value} - FPTP"
        elif "PAYMENT" in action_status:
            remark_text = "CURED - CONFIRM VIA SELECTIVE LIST"
        else:
            remark_text = ""
        
        dest_ws.cell(row=row, column=7).value = remark_text
        
        if "PAYMENT" in action_status:
            dest_ws.cell(row=row, column=9).value = ""
        else:
            dest_ws.cell(row=row, column=9).value = source_amount
        
        if "PAYMENT" in action_status:
            dest_ws.cell(row=row, column=10).value = source_amount
        else:
            dest_ws.cell(row=row, column=10).value = ""
        
        dest_ws.cell(row=row, column=11).value = source_collector
        
        if "PAYMENT" in action_status:
            dest_ws.cell(row=row, column=12).value = ""
        else:
            if source_phone1 and source_phone1 != "":
                dest_ws.cell(row=row, column=12).value = source_phone1
            else:
                dest_ws.cell(row=row, column=12).value = source_phone2
        
        if "PAYMENT" in action_status and source_date:
            if isinstance(source_date, datetime):
                formatted_paid_date = source_date.strftime("%m/%d/%Y")
            elif isinstance(source_date, date):
                formatted_paid_date = source_date.strftime("%m/%d/%Y")
            else:
                try:
                    date_obj = datetime.strptime(str(source_date), "%Y-%m-%d %H:%M:%S")
                    formatted_paid_date = date_obj.strftime("%m/%d/%Y")
                except:
                    try:
                        date_obj = datetime.strptime(str(source_date), "%Y-%m-%d")
                        formatted_paid_date = date_obj.strftime("%m/%d/%Y")
                    except:
                        formatted_paid_date = ""
            dest_ws.cell(row=row, column=14).value = formatted_paid_date
        else:
            dest_ws.cell(row=row, column=14).value = ""
    
    for row in range(2, final_row_count + 1):
        action_status = dest_ws.cell(row=row, column=2).value
        phone_no = dest_ws.cell(row=row, column=12).value
        
        if "PTP NEW" in action_status and phone_no:
            dest_ws.cell(row=row, column=7).value = f"1_{phone_no} - PTP NEW"
        elif "PTP FF" in action_status and phone_no:
            dest_ws.cell(row=row, column=7).value = f"{phone_no} - FPTP"
    
    for column in dest_ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        dest_ws.column_dimensions[column_letter].width = adjusted_width

    for row_idx in range(2, dest_ws.max_row + 1):  
        for col_idx in [3, 4, 14]: 
            cell = dest_ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_value_str = str(cell.value)
                cell.value = cell_value_str
                cell.number_format = '@'

    print(f"Saving Remarks file: {remarks_path}")
    dest_wb.save(remarks_path)
    
    print("Creating Others file...")
    others_wb = openpyxl.Workbook()
    others_ws = others_wb.active
    
    others_ws.cell(row=1, column=1).value = ws.cell(row=1, column=1).value 
    others_ws.cell(row=1, column=2).value = "REMARK BY" 
    
    for row in range(2, last_row + 1):
        others_ws.cell(row=row, column=1).value = ws.cell(row=row, column=1).value

        reference_value = ws.cell(row=row, column=1).value 
        
        for cured_row in range(2, ws.max_row + 1):  
            if ws.cell(row=cured_row, column=1).value == reference_value: 
                others_ws.cell(row=row, column=2).value = ws.cell(row=cured_row, column=2).value 
                break

    print(f"Saving Others file: {others_path}")
    others_wb.save(others_path)
    
    print("Creating Payments file...")
    payments_wb = openpyxl.Workbook()
    payments_ws = payments_wb.active
    payments_ws.cell(row=1, column=1).value = "LAN"
    payments_ws.cell(row=1, column=2).value = "ACCOUNT NUMBER"
    payments_ws.cell(row=1, column=3).value = "NAME"
    payments_ws.cell(row=1, column=4).value = "CARD NUMBER"
    payments_ws.cell(row=1, column=5).value = "PAYMENT AMOUNT"
    payments_ws.cell(row=1, column=6).value = "PAYMENT DATE"
    
    for row in range(2, last_row + 1):
        payments_ws.cell(row=row, column=1).value = ws.cell(row=row, column=17).value if ws.cell(row=row, column=17).value else ""
        payments_ws.cell(row=row, column=3).value = ws.cell(row=row, column=18).value if ws.cell(row=row, column=18).value else ""
        payments_ws.cell(row=row, column=5).value = ws.cell(row=row, column=4).value if ws.cell(row=row, column=4).value else ""
        date_value = ws.cell(row=row, column=3).value
        if date_value:
            if isinstance(date_value, datetime):
                formatted_date = date_value.strftime("%m/%d/%Y")
            else:
                formatted_date = str(date_value)
            payments_ws.cell(row=row, column=6).value = formatted_date
    
    for row in range(2, last_row + 1):
        payments_ws.cell(row=row, column=6).number_format = "@"
    
    for column in payments_ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        payments_ws.column_dimensions[column_letter].width = adjusted_width
    
    payments_wb.save(payments_path)
    print(f"Saving Payments file: {payments_path}")
    print("Total rows of payments and reshuffle: ", last_row - 1)
    print("All processing completed successfully.")

if __name__ == "__main__":
    import_data()